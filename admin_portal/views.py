import io

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Avg, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.utils.timezone import localdate, localtime
from django.views.decorators.http import require_http_methods
from functools import wraps

from core.models import Student, Teacher, User, StudentTeacherLink
from attendance.models import StudentAttendanceRecord, TeacherAttendanceRecord
from .forms import StudentForm, TeacherForm, SupervisorForm
from .models import AuditLog


def _log_audit(request, action, object_type, object_repr):
    """Create an AuditLog entry for a delete or edit action."""
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    ip = forwarded.split(',')[0].strip(
    ) if forwarded else request.META.get('REMOTE_ADDR', '')
    AuditLog.objects.create(
        action=action,
        actor_phone=request.user.phone,
        ip_address=ip,
        object_type=object_type,
        object_repr=str(object_repr)[:255],
    )


def _save_return(request, key):
    """Save the current full URL (with query string) to session under key."""
    request.session[key] = request.get_full_path()


def _get_return(request, key, fallback_url):
    """Return the saved list URL from session, or fallback_url."""
    url = request.session.get(key, '')
    # Safety: only use internal relative paths
    if url and url.startswith('/'):
        return url
    return fallback_url


def admin_required(view_func):
    """Decorator to ensure user is an admin."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_admin:
            messages.error(request, 'ليس لديك صلاحية الوصول لهذه الصفحة')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@admin_required
def dashboard(request):
    """Admin dashboard with summary statistics."""
    today = localdate()

    today_student_attendance = StudentAttendanceRecord.objects.filter(
        date=today)
    today_teacher_attendance = TeacherAttendanceRecord.objects.filter(
        date=today)

    context = {
        'total_students': Student.objects.count(),
        'total_teachers': Teacher.objects.count(),
        'total_users': User.objects.count(),
        'today_student_attendance_count': today_student_attendance.count(),
        'today_teacher_attendance_count': today_teacher_attendance.count(),
        'today_substitute_count': today_student_attendance.filter(
            original_teacher__isnull=False,
            assigned_teacher__isnull=False,
        ).exclude(original_teacher=models.F('assigned_teacher')).count(),
        'today_missing_photos_count': today_student_attendance.filter(
            models.Q(daily_photo__isnull=True) | models.Q(daily_photo='')
        ).count(),
        'today': today,
    }
    return render(request, 'admin_portal/dashboard.html', context)


# ---------------------------------------------------------------------------
# Student management
# ---------------------------------------------------------------------------

@admin_required
def student_list(request):
    """Paginated student list with search and grade filter."""
    _save_return(request, 'student_list_return')
    q = request.GET.get('q', '').strip()
    grade_filter = request.GET.get('grade', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    birth_year_filter = request.GET.get('birth_year', '').strip()
    hall_filter = request.GET.get('hall', '').strip()
    sort = request.GET.get('sort', '').strip()

    qs = Student.objects.annotate(avg_rating=Avg('attendance_records__rating'))
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(national_id__icontains=q)
            | Q(student_code__icontains=q)
            | Q(nickname__icontains=q)
            | Q(phone__icontains=q)
            | Q(parent_phone__icontains=q)
        )
    if grade_filter:
        qs = qs.filter(grade=grade_filter)
    if gender_filter:
        qs = qs.filter(gender=gender_filter)
    if birth_year_filter:
        qs = qs.filter(date_of_birth__year=birth_year_filter)
    if hall_filter:
        qs = qs.filter(hall_name=hall_filter)

    if sort == 'avg_rating_desc':
        qs = qs.order_by(models.F('avg_rating').desc(nulls_last=True))
    elif sort == 'avg_rating_asc':
        qs = qs.order_by(models.F('avg_rating').asc(nulls_last=True))
    elif sort == 'name_asc':
        qs = qs.order_by('full_name')
    elif sort == 'name_desc':
        qs = qs.order_by('-full_name')

    grades = (
        Student.objects
        .exclude(grade__isnull=True)
        .exclude(grade='')
        .values_list('grade', flat=True)
        .distinct()
        .order_by('grade')
    )

    birth_years = (
        Student.objects
        .exclude(date_of_birth__isnull=True)
        .values_list('date_of_birth__year', flat=True)
        .distinct()
        .order_by('-date_of_birth__year')
    )

    halls = (
        Student.objects
        .exclude(hall_name='').exclude(hall_name__isnull=True)
        .values_list('hall_name', flat=True)
        .distinct()
        .order_by('hall_name')
    )

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_portal/students.html', {
        'page_obj': page_obj,
        'q': q,
        'grade_filter': grade_filter,
        'gender_filter': gender_filter,
        'birth_year_filter': birth_year_filter,
        'hall_filter': hall_filter,
        'grades': grades,
        'birth_years': birth_years,
        'halls': halls,
        'total_count': qs.count(),
        'sort': sort,
    })


@admin_required
def student_create(request):
    """Create a new student."""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(
                request, f'تم إضافة الطالب "{student.full_name}" بنجاح')
            return redirect(_get_return(request, 'student_list_return',
                                        reverse('admin_portal:student_list')))
    else:
        form = StudentForm()

    return render(request, 'admin_portal/student_form.html', {
        'form': form,
        'title': 'إضافة طالب جديد',
        'submit_label': 'إضافة',
    })


@admin_required
def student_edit(request, pk):
    """Edit an existing student."""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            _log_audit(request, AuditLog.Action.EDIT,
                       'طالب', student.full_name)
            messages.success(
                request, f'تم تحديث بيانات "{student.full_name}" بنجاح')
            return redirect(_get_return(request, 'student_list_return',
                                        reverse('admin_portal:student_list')))
    else:
        form = StudentForm(instance=student)

    return render(request, 'admin_portal/student_form.html', {
        'form': form,
        'student': student,
        'title': f'تعديل: {student.full_name}',
        'submit_label': 'حفظ التغييرات',
    })


@admin_required
@require_http_methods(['POST'])
def student_delete(request, pk):
    """Delete a student (POST only)."""
    student = get_object_or_404(Student, pk=pk)
    name = student.full_name
    _log_audit(request, AuditLog.Action.DELETE, 'طالب', name)
    student.delete()
    messages.success(request, f'تم حذف الطالب "{name}" بنجاح')
    return redirect(_get_return(request, 'student_list_return',
                                reverse('admin_portal:student_list')))


@admin_required
@require_http_methods(['POST'])
def student_bulk_delete(request):
    """Bulk-delete selected students (POST only)."""
    ids = request.POST.getlist('student_ids')
    if not ids:
        messages.warning(request, 'لم يتم تحديد أي طالب')
        return redirect('admin_portal:student_list')
    qs = Student.objects.filter(pk__in=ids)
    names = list(qs.values_list('full_name', flat=True))
    count = len(names)
    for name in names:
        _log_audit(request, AuditLog.Action.DELETE, 'طالب', name)
    qs.delete()
    messages.success(request, f'تم حذف {count} طالب بنجاح')
    return redirect(_get_return(request, 'student_list_return',
                                reverse('admin_portal:student_list')))


@admin_required
def student_detail(request, pk):
    """Full profile page for a single student."""
    student = get_object_or_404(Student, pk=pk)
    teachers = (
        StudentTeacherLink.objects
        .filter(student=student)
        .select_related('teacher')
    )
    from attendance.models import StudentAttendanceRecord
    recent_records = (
        StudentAttendanceRecord.objects
        .filter(student=student)
        .select_related('assigned_teacher')
        .order_by('-date')[:10]
    )
    total_records = StudentAttendanceRecord.objects.filter(
        student=student).count()
    return render(request, 'admin_portal/student_detail.html', {
        'student': student,
        'teachers': teachers,
        'recent_records': recent_records,
        'total_records': total_records,
    })


@admin_required
@require_http_methods(['POST'])
def student_import(request):
    """Bulk-import students from an uploaded Excel file."""
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'الرجاء اختيار ملف Excel')
        return redirect('admin_portal:student_list')

    if not excel_file.name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        messages.error(request, 'الملف يجب أن يكون بصيغة Excel (.xlsx)')
        return redirect('admin_portal:student_list')

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        messages.error(request, 'تعذّر قراءة الملف. تأكد أنه ملف Excel صالح.')
        return redirect('admin_portal:student_list')

    if len(rows) < 2:
        messages.warning(request, 'الملف لا يحتوي على بيانات')
        return redirect('admin_portal:student_list')

    # Normalise header row
    headers = [str(h).strip().lower()
               if h is not None else '' for h in rows[0]]
    if not {'full_name', 'national_id'}.issubset(set(headers)):
        messages.error(
            request, 'يجب أن يحتوي الملف على أعمدة: full_name و national_id')
        return redirect('admin_portal:student_list')

    col = {h: i for i, h in enumerate(headers) if h}

    created = skipped = 0
    duplicate_rows = []
    error_msgs = []

    for row_num, row in enumerate(rows[1:], start=2):
        def _cell(name, _row=row):
            idx = col.get(name)
            if idx is None or idx >= len(_row):
                return ''
            return str(_row[idx] or '').strip()

        full_name = _cell('full_name')
        national_id = _cell('national_id')

        # Skip entirely blank rows
        if not full_name and not national_id:
            continue

        if not full_name:
            error_msgs.append(f'الصف {row_num}: الاسم الكامل مطلوب')
            continue
        if not national_id:
            error_msgs.append(f'الصف {row_num}: الرقم القومي مطلوب')
            continue

        if Student.objects.filter(national_id=national_id).exists():
            skipped += 1
            duplicate_rows.append(str(row_num))
            continue

        student_code = _cell('student_code') or None
        grade = _cell('grade') or None
        phone = _cell('phone') or None
        parent_phone = _cell('parent_phone') or None
        gender_val = _cell('gender') or None
        if gender_val and gender_val.upper() in ('M', 'F'):
            gender_val = gender_val.upper()
        else:
            gender_val = None

        # New fields
        nickname = _cell('nickname') or ''
        hall_name = _cell('hall_name') or ''
        notes = _cell('notes') or ''
        parent_full_name = _cell('parent_full_name') or ''
        parent_qualification = _cell('parent_qualification') or ''
        parent_job = _cell('parent_job') or ''
        parent_calls_phone = _cell('parent_calls_phone') or None
        parent_spouse_job = _cell('parent_spouse_job') or ''
        parent_address = _cell('parent_address') or ''
        child_pickup_person = _cell('child_pickup_person') or ''

        marital_raw = _cell('parent_marital_status').lower()
        valid_marital = {'married', 'divorced', 'widowed', 'separated'}
        parent_marital_status = marital_raw if marital_raw in valid_marital else ''

        def _parse_date(val):
            if not val:
                return None
            from datetime import date as _date, datetime as _dt
            if isinstance(val, _dt):
                return val.date()
            if isinstance(val, _date):
                return val
            s = str(val).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return _dt.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        date_of_birth = _parse_date(
            row[col['date_of_birth']] if 'date_of_birth' in col else None)
        joining_date = _parse_date(
            row[col['joining_date']] if 'joining_date' in col else None)

        try:
            Student.objects.create(
                full_name=full_name,
                national_id=national_id,
                student_code=student_code,
                grade=grade,
                phone=phone,
                parent_phone=parent_phone,
                gender=gender_val,
                nickname=nickname,
                hall_name=hall_name,
                notes=notes,
                date_of_birth=date_of_birth,
                joining_date=joining_date,
                parent_full_name=parent_full_name,
                parent_qualification=parent_qualification,
                parent_job=parent_job,
                parent_calls_phone=parent_calls_phone,
                parent_marital_status=parent_marital_status,
                parent_spouse_job=parent_spouse_job,
                parent_address=parent_address,
                child_pickup_person=child_pickup_person,
            )
            created += 1
        except Exception as exc:
            error_msgs.append(f'الصف {row_num}: {exc}')

    # Feedback messages
    if created:
        messages.success(request, f'تمت إضافة {created} طالب بنجاح')
    if skipped:
        rows_preview = ', '.join(duplicate_rows[:10])
        if len(duplicate_rows) > 10:
            rows_preview += f' ... (+{len(duplicate_rows) - 10})'
        messages.warning(
            request,
            f'تم تخطي {skipped} سجل مكرر (الرقم القومي موجود مسبقاً). الصفوف: {rows_preview}'
        )
    if error_msgs:
        preview = ' | '.join(error_msgs[:5])
        if len(error_msgs) > 5:
            preview += f' ... (+{len(error_msgs) - 5} أخطاء أخرى)'
        messages.error(
            request, f'{len(error_msgs)} خطأ أثناء الاستيراد: {preview}')
    if not created and not skipped and not error_msgs:
        messages.info(request, 'لم يتم العثور على بيانات جديدة للاستيراد')

    return redirect('admin_portal:student_list')


@admin_required
@require_http_methods(['POST'])
def teacher_import(request):
    """Bulk-import teachers from an uploaded Excel file."""
    from django.db import transaction

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'الرجاء اختيار ملف Excel')
        return redirect('admin_portal:teacher_list')

    if not excel_file.name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        messages.error(request, 'الملف يجب أن يكون بصيغة Excel (.xlsx)')
        return redirect('admin_portal:teacher_list')

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        messages.error(request, 'تعذّر قراءة الملف. تأكد أنه ملف Excel صالح.')
        return redirect('admin_portal:teacher_list')

    if len(rows) < 2:
        messages.warning(request, 'الملف لا يحتوي على بيانات')
        return redirect('admin_portal:teacher_list')

    headers = [str(h).strip().lower()
               if h is not None else '' for h in rows[0]]
    required = {'full_name', 'phone', 'password'}
    if not required.issubset(set(headers)):
        missing = required - set(headers)
        messages.error(
            request,
            f'يجب أن يحتوي الملف على الأعمدة: {", ".join(missing)}')
        return redirect('admin_portal:teacher_list')

    col = {h: i for i, h in enumerate(headers) if h}

    created = skipped = 0
    error_msgs = []

    for row_num, row in enumerate(rows[1:], start=2):
        def _cell(name, _row=row):
            idx = col.get(name)
            if idx is None or idx >= len(_row):
                return ''
            return str(_row[idx] or '').strip()

        full_name = _cell('full_name')
        phone = _cell('phone')
        password = _cell('password')

        if not full_name and not phone:
            continue  # blank row

        if not full_name:
            error_msgs.append(f'الصف {row_num}: الاسم الكامل مطلوب')
            continue
        if not phone:
            error_msgs.append(f'الصف {row_num}: رقم الهاتف مطلوب')
            continue
        if not password:
            error_msgs.append(f'الصف {row_num}: كلمة المرور مطلوبة')
            continue

        import re as _re
        if not _re.match(r'^0\d{10}$', phone):
            error_msgs.append(f'الصف {row_num}: رقم الهاتف غير صالح ({phone})')
            continue

        if User.objects.filter(phone=phone).exists():
            skipped += 1
            continue

        subject = _cell('subject') or None
        first_name = _cell('first_name') or ''
        last_name = _cell('last_name') or ''
        gender_val = _cell('gender') or None
        if gender_val and gender_val.upper() in ('M', 'F'):
            gender_val = gender_val.upper()
        else:
            gender_val = None

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    phone=phone,
                    password=password,
                    role=User.Role.TEACHER,
                    first_name=first_name,
                    last_name=last_name,
                )
                Teacher.objects.create(
                    user=user,
                    full_name=full_name,
                    subject=subject,
                    gender=gender_val,
                )
            created += 1
        except Exception as exc:
            error_msgs.append(f'الصف {row_num}: {exc}')

    if created:
        messages.success(request, f'تمت إضافة {created} معلم بنجاح')
    if skipped:
        messages.warning(
            request, f'تم تخطي {skipped} سجل مكرر (رقم الهاتف مستخدم مسبقاً)')
    if error_msgs:
        preview = ' | '.join(error_msgs[:5])
        if len(error_msgs) > 5:
            preview += f' ... (+{len(error_msgs) - 5} أخطاء أخرى)'
        messages.error(
            request, f'{len(error_msgs)} خطأ أثناء الاستيراد: {preview}')
    if not created and not skipped and not error_msgs:
        messages.info(request, 'لم يتم العثور على بيانات جديدة للاستيراد')

    return redirect('admin_portal:teacher_list')


@admin_required
@require_http_methods(['GET'])
def teacher_import_template(request):
    """Return a blank Excel template for bulk teacher import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Teachers'
    ws.append(['full_name', 'phone', 'password', 'subject',
              'first_name', 'last_name', 'gender'])
    ws.append(['أحمد محمد', '01012345678', 'password123',
              'رياضيات', 'أحمد', 'محمد', 'M'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="teachers_import_template.xlsx"'
    return response


@admin_required
@require_http_methods(['GET'])
def student_import_template(request):
    """Return a blank Excel template for bulk student import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append([
        'full_name', 'national_id', 'student_code', 'grade', 'gender', 'phone',
        'nickname', 'date_of_birth', 'joining_date', 'hall_name', 'notes',
        'parent_phone', 'parent_full_name', 'parent_qualification', 'parent_job',
        'parent_calls_phone', 'parent_marital_status', 'parent_spouse_job',
        'parent_address', 'child_pickup_person',
    ])
    ws.append([
        'أحمد محمد علي', '12345678901234', 'STU001', 'السنة الأولى', 'M', '',
        'أحمد', '2010-05-15', '2024-09-01', 'قاعة 1', '',
        '01012345678', 'محمد علي حسن سالم', 'بكالوريوس', 'مهندس',
        '01098765432', 'married', '', 'القاهرة', '',
    ])
    # Add a note row explaining parent_marital_status valid values
    ws.append([])
    ws.append(
        ['# parent_marital_status القيم المقبولة: married / divorced / widowed / separated'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'
    return response


@admin_required
@require_http_methods(['GET'])
def student_export(request):
    """Export all student records (with parent data) to Excel."""
    q = request.GET.get('q', '').strip()
    grade_filter = request.GET.get('grade', '').strip()
    gender_filter = request.GET.get('gender', '').strip()

    qs = Student.objects.all().order_by('full_name')
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(national_id__icontains=q)
            | Q(student_code__icontains=q)
            | Q(nickname__icontains=q)
        )
    if grade_filter:
        qs = qs.filter(grade=grade_filter)
    if gender_filter:
        qs = qs.filter(gender=gender_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append([
        'full_name', 'national_id', 'student_code', 'grade', 'gender', 'phone',
        'nickname', 'date_of_birth', 'joining_date', 'hall_name', 'notes',
        'parent_phone', 'parent_full_name', 'parent_qualification', 'parent_job',
        'parent_calls_phone', 'parent_marital_status', 'parent_spouse_job',
        'parent_address', 'child_pickup_person',
    ])
    for s in qs:
        ws.append([
            s.full_name,
            s.national_id,
            s.student_code or '',
            s.grade or '',
            s.gender or '',
            s.phone or '',
            s.nickname,
            s.date_of_birth.isoformat() if s.date_of_birth else '',
            s.joining_date.isoformat() if s.joining_date else '',
            s.hall_name,
            s.notes,
            s.parent_phone or '',
            s.parent_full_name,
            s.parent_qualification,
            s.parent_job,
            s.parent_calls_phone or '',
            s.parent_marital_status,
            s.parent_spouse_job,
            s.parent_address,
            s.child_pickup_person,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
    return response


# ---------------------------------------------------------------------------
# Teacher management
# ---------------------------------------------------------------------------

@admin_required
def teacher_list(request):
    """Paginated teacher list with name / phone / subject search."""
    _save_return(request, 'teacher_list_return')
    q = request.GET.get('q', '').strip()
    gender_filter = request.GET.get('gender', '').strip()

    qs = Teacher.objects.select_related('user').all()
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(subject__icontains=q)
            | Q(user__phone__icontains=q)
        )
    if gender_filter:
        qs = qs.filter(gender=gender_filter)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_portal/teachers.html', {
        'page_obj': page_obj,
        'q': q,
        'gender_filter': gender_filter,
        'total_count': qs.count(),
    })


@admin_required
def teacher_create(request):
    """Create a new teacher with a linked user account."""
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            teacher = form.save()
            messages.success(
                request, f'تم إضافة المعلم "{teacher.full_name}" بنجاح')
            return redirect(_get_return(request, 'teacher_list_return',
                                        reverse('admin_portal:teacher_list')))
    else:
        form = TeacherForm()

    return render(request, 'admin_portal/teacher_form.html', {
        'form': form,
        'title': 'إضافة معلم جديد',
        'submit_label': 'إضافة',
    })


@admin_required
def teacher_edit(request, pk):
    """Edit an existing teacher and their linked user account."""
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=pk)

    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            _log_audit(request, AuditLog.Action.EDIT,
                       'معلم', teacher.full_name)
            messages.success(
                request, f'تم تحديث بيانات "{teacher.full_name}" بنجاح')
            return redirect(_get_return(request, 'teacher_list_return',
                                        reverse('admin_portal:teacher_list')))
    else:
        form = TeacherForm(
            initial={
                'full_name': teacher.full_name,
                'subject': teacher.subject or '',
                'phone': teacher.user.phone,
                'first_name': teacher.user.first_name,
                'last_name': teacher.user.last_name,
            },
            instance=teacher,
        )

    return render(request, 'admin_portal/teacher_form.html', {
        'form': form,
        'teacher': teacher,
        'title': f'تعديل: {teacher.full_name}',
        'submit_label': 'حفظ التغييرات',
    })


@admin_required
@require_http_methods(['POST'])
def teacher_delete(request, pk):
    """Delete a teacher (and their user account) — POST only."""
    teacher = get_object_or_404(Teacher, pk=pk)
    name = teacher.full_name
    _log_audit(request, AuditLog.Action.DELETE, 'معلم', name)
    # Deleting the user cascades to the Teacher profile
    teacher.user.delete()
    messages.success(request, f'تم حذف المعلم "{name}" بنجاح')
    return redirect(_get_return(request, 'teacher_list_return',
                                reverse('admin_portal:teacher_list')))


# ---------------------------------------------------------------------------
# Student-Teacher linking
# ---------------------------------------------------------------------------

@admin_required
def teacher_students(request, pk):
    """Manage which students are linked to a teacher.

    GET  – renders all students with checkboxes (linked ones pre-ticked).
    POST – atomically replaces the teacher's student links with the
           submitted set of student IDs and primary flags.
    """
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=pk)

    # Search/filter within the student pool
    q = request.GET.get('q', '').strip()
    grade_filter = request.GET.get('grade', '').strip()

    all_students = Student.objects.all()
    if q:
        all_students = all_students.filter(
            Q(full_name__icontains=q)
            | Q(national_id__icontains=q)
            | Q(student_code__icontains=q)
        )
    if grade_filter:
        all_students = all_students.filter(grade=grade_filter)

    grades = (
        Student.objects
        .exclude(grade__isnull=True).exclude(grade='')
        .values_list('grade', flat=True).distinct().order_by('grade')
    )

    if request.method == 'POST':
        # Collect submitted student IDs and which ones are marked primary
        selected_ids = set(request.POST.getlist('students'))
        primary_ids = set(request.POST.getlist('primary'))

        # Validate that every submitted ID is a real Student UUID
        existing_ids = set(
            Student.objects.filter(pk__in=selected_ids)
            .values_list('id', flat=True)
        )
        # Cast to str for comparison (UUIDs come back as UUID objects)
        existing_ids_str = {str(i) for i in existing_ids}
        invalid = selected_ids - existing_ids_str
        if invalid:
            messages.error(request, 'بعض الطلاب المحددين غير موجودين')
            return redirect('admin_portal:teacher_students', pk=pk)

        from django.db import transaction
        with transaction.atomic():
            # Remove links for de-selected students
            StudentTeacherLink.objects.filter(teacher=teacher).exclude(
                student_id__in=existing_ids
            ).delete()

            # Upsert links for selected students
            for sid in existing_ids_str:
                is_primary = sid in primary_ids
                StudentTeacherLink.objects.update_or_create(
                    teacher=teacher,
                    student_id=sid,
                    defaults={'is_primary': is_primary},
                )

        count = len(existing_ids_str)
        messages.success(
            request, f'تم تحديث قائمة طلاب "{teacher.full_name}" — {count} طالب مرتبط')
        return redirect('admin_portal:teacher_students', pk=pk)

    # Build sets of currently-linked and primary student IDs for template use
    links_qs = StudentTeacherLink.objects.filter(teacher=teacher)
    linked_ids = {str(link.student_id) for link in links_qs}
    primary_ids = {str(link.student_id)
                   for link in links_qs if link.is_primary}

    return render(request, 'admin_portal/teacher_students.html', {
        'teacher': teacher,
        'students': all_students,
        'linked_ids': linked_ids,
        'primary_ids': primary_ids,
        'q': q,
        'grade_filter': grade_filter,
        'grades': grades,
        'linked_count': len(linked_ids),
    })


@admin_required
def teacher_students_export(request, pk):
    """Export the linked students of a teacher to Excel."""
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=pk)

    linked_students = (
        Student.objects
        .filter(teacher_links__teacher=teacher)
        .select_related()
        .order_by('full_name')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'طلاب المعلم'
    ws.append([
        'full_name', 'national_id', 'student_code', 'grade', 'gender', 'phone',
        'nickname', 'date_of_birth', 'joining_date', 'hall_name', 'notes',
        'parent_phone', 'parent_full_name', 'parent_qualification', 'parent_job',
        'parent_calls_phone', 'parent_marital_status', 'parent_spouse_job',
        'parent_address', 'child_pickup_person',
    ])
    for s in linked_students:
        ws.append([
            s.full_name,
            s.national_id or '',
            s.student_code or '',
            s.grade or '',
            s.gender or '',
            s.phone or '',
            s.nickname,
            s.date_of_birth.isoformat() if s.date_of_birth else '',
            s.joining_date.isoformat() if s.joining_date else '',
            s.hall_name,
            s.notes,
            s.parent_phone or '',
            s.parent_full_name,
            s.parent_qualification,
            s.parent_job,
            s.parent_calls_phone or '',
            s.parent_marital_status,
            s.parent_spouse_job,
            s.parent_address,
            s.child_pickup_person,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = teacher.full_name.replace(' ', '_')
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="students_{safe_name}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# Supervisor management
# ---------------------------------------------------------------------------

@admin_required
def supervisor_list(request):
    """List all supervisor accounts."""
    _save_return(request, 'supervisor_list_return')
    q = request.GET.get('q', '').strip()
    qs = User.objects.filter(role=User.Role.SUPERVISOR).order_by('first_name')
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) | Q(phone__icontains=q)
        )
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'admin_portal/supervisors.html', {
        'page_obj': page_obj,
        'q': q,
        'total_count': qs.count(),
    })


@admin_required
def supervisor_create(request):
    """Create a new supervisor account."""
    if request.method == 'POST':
        form = SupervisorForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(
                request, f'تم إضافة المشرف "{user.first_name}" بنجاح')
            return redirect(_get_return(request, 'supervisor_list_return',
                                        reverse('admin_portal:supervisor_list')))
    else:
        form = SupervisorForm()
    return render(request, 'admin_portal/supervisor_form.html', {
        'form': form,
        'title': 'إضافة مشرف جديد',
        'submit_label': 'إضافة',
    })


@admin_required
def supervisor_edit(request, pk):
    """Edit an existing supervisor account."""
    supervisor = get_object_or_404(User, pk=pk, role=User.Role.SUPERVISOR)
    if request.method == 'POST':
        form = SupervisorForm(request.POST, instance=supervisor)
        if form.is_valid():
            form.save()
            _log_audit(request, AuditLog.Action.EDIT,
                       'مشرف', supervisor.first_name)
            messages.success(
                request, f'تم تحديث بيانات "{supervisor.first_name}" بنجاح')
            return redirect(_get_return(request, 'supervisor_list_return',
                                        reverse('admin_portal:supervisor_list')))
    else:
        form = SupervisorForm(
            initial={'full_name': supervisor.first_name,
                     'phone': supervisor.phone},
            instance=supervisor,
        )
    return render(request, 'admin_portal/supervisor_form.html', {
        'form': form,
        'supervisor': supervisor,
        'title': f'تعديل: {supervisor.first_name}',
        'submit_label': 'حفظ التغييرات',
    })


@admin_required
@require_http_methods(['POST'])
def supervisor_delete(request, pk):
    """Delete a supervisor account (POST only)."""
    supervisor = get_object_or_404(User, pk=pk, role=User.Role.SUPERVISOR)
    name = supervisor.first_name
    _log_audit(request, AuditLog.Action.DELETE, 'مشرف', name)
    supervisor.delete()
    messages.success(request, f'تم حذف المشرف "{name}" بنجاح')
    return redirect(_get_return(request, 'supervisor_list_return',
                                reverse('admin_portal:supervisor_list')))


# ---------------------------------------------------------------------------
# Student attendance history
# ---------------------------------------------------------------------------

@admin_required
def student_history(request, pk):
    """Full attendance history for a single student."""
    student = get_object_or_404(Student, pk=pk)
    records = (
        StudentAttendanceRecord.objects
        .filter(student=student)
        .select_related('assigned_teacher', 'original_teacher')
        .order_by('-date', '-check_in_time')
    )
    teachers = (
        StudentTeacherLink.objects
        .filter(student=student)
        .select_related('teacher')
    )
    total_records = records.count()
    avg_rating = records.aggregate(avg=Avg('rating'))['avg']
    oldest_record = records.last() if total_records else None
    return render(request, 'admin_portal/student_history.html', {
        'student': student,
        'records': records,
        'total_records': total_records,
        'teachers': teachers,
        'oldest_record': oldest_record,
        'avg_rating': avg_rating,
    })


# ---------------------------------------------------------------------------
# Attendance records browser
# ---------------------------------------------------------------------------

@admin_required
def attendance_records(request):
    """Filterable table of student and teacher attendance records (tabbed)."""
    tab = request.GET.get('tab', 'students')  # 'students' | 'teachers'

    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    teacher_id = request.GET.get('teacher', '').strip()
    student_q = request.GET.get('student', '').strip()
    grade = request.GET.get('grade', '').strip()
    teacher_q = request.GET.get('teacher_q', '').strip()

    # --- student records ---
    student_qs = (
        StudentAttendanceRecord.objects
        .select_related('student', 'assigned_teacher', 'original_teacher')
        .order_by('-date', '-check_in_time')
    )
    if date_from:
        student_qs = student_qs.filter(date__gte=date_from)
    if date_to:
        student_qs = student_qs.filter(date__lte=date_to)
    if teacher_id:
        student_qs = student_qs.filter(
            Q(assigned_teacher_id=teacher_id) | Q(
                original_teacher_id=teacher_id)
        )
    if student_q:
        student_qs = student_qs.filter(
            Q(student__full_name__icontains=student_q)
            | Q(student__national_id__icontains=student_q)
            | Q(student__student_code__icontains=student_q)
        )
    if grade:
        student_qs = student_qs.filter(student__grade=grade)

    # --- teacher records ---
    teacher_qs = (
        TeacherAttendanceRecord.objects
        .select_related('teacher', 'teacher__user')
        .order_by('-date', '-check_in_time')
    )
    if date_from:
        teacher_qs = teacher_qs.filter(date__gte=date_from)
    if date_to:
        teacher_qs = teacher_qs.filter(date__lte=date_to)
    if teacher_q:
        teacher_qs = teacher_qs.filter(
            Q(teacher__full_name__icontains=teacher_q)
            | Q(teacher__user__phone__icontains=teacher_q)
        )

    student_total = student_qs.count()
    teacher_total = teacher_qs.count()

    if tab == 'teachers':
        paginator = Paginator(teacher_qs, 50)
        teacher_page = paginator.get_page(request.GET.get('page'))
        student_page = None
    else:
        paginator = Paginator(student_qs, 50)
        student_page = paginator.get_page(request.GET.get('page'))
        teacher_page = None

    teachers = Teacher.objects.select_related('user').order_by('full_name')
    grades = (
        Student.objects
        .exclude(grade__isnull=True).exclude(grade='')
        .values_list('grade', flat=True).distinct().order_by('grade')
    )

    return render(request, 'admin_portal/attendance_records.html', {
        'tab': tab,
        'student_page': student_page,
        'teacher_page': teacher_page,
        'student_total': student_total,
        'teacher_total': teacher_total,
        'teachers': teachers,
        'grades': grades,
        'date_from': date_from,
        'date_to': date_to,
        'teacher_id': teacher_id,
        'student_q': student_q,
        'grade': grade,
        'teacher_q': teacher_q,
    })


# ---------------------------------------------------------------------------
# Excel export for attendance records
# ---------------------------------------------------------------------------

@admin_required
def export_attendance_excel(request):
    """Download filtered attendance records as .xlsx.

    Accepts the same GET params as attendance_records.
    """
    tab = request.GET.get('tab', 'students')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    teacher_id = request.GET.get('teacher', '').strip()
    student_q = request.GET.get('student', '').strip()
    grade = request.GET.get('grade', '').strip()
    teacher_q = request.GET.get('teacher_q', '').strip()

    wb = openpyxl.Workbook()
    ws = wb.active

    format_type = request.GET.get('format', 'old')

    if tab == 'teachers' and format_type == 'pivot':
        import datetime
        from openpyxl.styles import Alignment

        ws.title = 'حضور المعلمين (مجمع)'
        
        # Parse date range
        d_from = None
        d_to = None
        if date_from:
            try:
                d_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_to:
            try:
                d_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                pass
                
        # Base querysets
        teachers_qs = Teacher.objects.all().order_by('full_name')
        if teacher_q:
            teachers_qs = teachers_qs.filter(
                Q(full_name__icontains=teacher_q)
                | Q(user__phone__icontains=teacher_q)
            )

        qs = TeacherAttendanceRecord.objects.filter(teacher__in=teachers_qs)
        if d_from:
            qs = qs.filter(date__gte=d_from)
        if d_to:
            qs = qs.filter(date__lte=d_to)
            
        # If dates not provided, find min/max in qs
        if not d_from and qs.exists():
            d_from = qs.aggregate(models.Min('date'))['date__min']
        if not d_to and qs.exists():
            d_to = qs.aggregate(models.Max('date'))['date__max']
            
        # Fallback to today
        if not d_from:
            d_from = localdate()
        if not d_to:
            d_to = localdate()
            
        # Build date list
        date_list = []
        curr = d_from
        while curr <= d_to and (curr - d_from).days < 366:
            date_list.append(curr)
            curr += datetime.timedelta(days=1)

        # Build attendance map
        attendance_map = {}
        for rec in qs.order_by('date', 'check_in_time'):
            attendance_map[(rec.teacher_id, rec.date)] = rec

        # Build Headers
        header1 = ['المعلم']
        header2 = ['']
        for d in date_list:
            header1.extend([str(d), '', ''])
            header2.extend(['الحضور', 'المغادرة', 'المدة'])
            
        ws.append(header1)
        ws.append(header2)
        
        # Merge date cells in header1 and center align
        for i, d in enumerate(date_list):
            start_col = 2 + (i * 3)
            end_col = start_col + 2
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Populate rows
        for t in teachers_qs:
            row = [t.full_name]
            for d in date_list:
                rec = attendance_map.get((t.id, d))
                if rec:
                    check_in = localtime(rec.check_in_time).strftime('%H:%M') if rec.check_in_time else ''
                    check_out = localtime(rec.check_out_time).strftime('%H:%M') if rec.check_out_time else ''
                    duration = rec.duration_display if rec.check_out_time else ''
                    row.extend([check_in, check_out, duration])
                else:
                    row.extend(['absent', '', ''])
            ws.append(row)

        filename = 'teacher_attendance_pivot'
    elif tab == 'teachers':
        ws.title = 'حضور المعلمين'
        qs = (
            TeacherAttendanceRecord.objects
            .select_related('teacher', 'teacher__user')
            .order_by('-date', '-check_in_time')
        )
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if teacher_q:
            qs = qs.filter(
                Q(teacher__full_name__icontains=teacher_q)
                | Q(teacher__user__phone__icontains=teacher_q)
            )
        ws.append(['التاريخ', 'المعلم', 'وقت الحضور',
                  'وقت المغادرة', 'مدة الحضور', 'التقييم', 'ملاحظات'])
        for rec in qs:
            ws.append([
                str(rec.date),
                rec.teacher.full_name,
                localtime(rec.check_in_time).strftime('%H:%M'),
                localtime(rec.check_out_time).strftime(
                    '%H:%M') if rec.check_out_time else '',
                rec.duration_display if rec.check_out_time else '',
                rec.rating,
                rec.notes,
            ])
        filename = 'teacher_attendance'
    else:
        ws.title = 'حضور الطلاب'
        qs = (
            StudentAttendanceRecord.objects
            .select_related('student', 'assigned_teacher', 'original_teacher')
            .order_by('-date', '-check_in_time')
        )
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if teacher_id:
            qs = qs.filter(
                Q(assigned_teacher_id=teacher_id) | Q(
                    original_teacher_id=teacher_id)
            )
        if student_q:
            qs = qs.filter(
                Q(student__full_name__icontains=student_q)
                | Q(student__national_id__icontains=student_q)
                | Q(student__student_code__icontains=student_q)
            )
        if grade:
            qs = qs.filter(student__grade=grade)
        ws.append([
            'التاريخ', 'اسم الطالب', 'الرقم القومي', 'الكود', 'الصف',
            'وقت الحضور', 'وقت المغادرة', 'مدة الحضور', 'المعلم المكلف', 'التقييم', 'نيابة', 'ملاحظات', 'ملاحظة المعلم',
        ])
        for rec in qs:
            ws.append([
                str(rec.date),
                rec.student.full_name,
                rec.student.national_id,
                rec.student.student_code or '',
                rec.student.grade or '',
                localtime(rec.check_in_time).strftime('%H:%M'),
                localtime(rec.check_out_time).strftime(
                    '%H:%M') if rec.check_out_time else '',
                rec.duration_display if rec.check_out_time else '',
                rec.assigned_teacher.full_name if rec.assigned_teacher else '',
                rec.rating,
                'نعم' if rec.is_substitute_assignment else 'لا',
                rec.substitute_note,
                rec.teacher_note,
            ])
        filename = 'student_attendance'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# Edit student attendance record rating (admin + linked teacher)
# ---------------------------------------------------------------------------

@login_required
def attendance_record_edit_rating(request, pk):
    """Allow admin or a teacher/supervisor linked to the student to edit the record rating."""
    record = get_object_or_404(
        StudentAttendanceRecord.objects.select_related('student'), pk=pk
    )
    user = request.user
    acting_as_teacher = False  # True when supervisor acts as a teacher
    if user.is_admin:
        can_edit = True
    elif user.is_teacher:
        can_edit = StudentTeacherLink.objects.filter(
            teacher__user=user, student=record.student
        ).exists()
    elif user.is_supervisor:
        teacher_pk = request.session.get('supervisor_teacher_id')
        if teacher_pk:
            can_edit = StudentTeacherLink.objects.filter(
                teacher_id=teacher_pk, student=record.student
            ).exists()
            acting_as_teacher = can_edit
        else:
            can_edit = False
    else:
        can_edit = False

    if not can_edit:
        messages.error(request, 'ليس لديك صلاحية تعديل هذا السجل')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            rating = int(request.POST.get('rating', ''))
            if not 1 <= rating <= 10:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'التقييم يجب أن يكون رقماً من 1 إلى 10')
        else:
            record.rating = rating
            record.save(update_fields=['rating'])
            _log_audit(request, AuditLog.Action.EDIT, 'سجل حضور طالب',
                       f'{record.student.full_name} — {record.date}')
            messages.success(request, f'تم تحديث التقييم إلى {rating}/10')
            if user.is_admin:
                return redirect('admin_portal:student_history', pk=record.student_id)
            return redirect('teacher_portal:dashboard')

    return render(request, 'admin_portal/attendance_record_edit_rating.html', {
        'record': record,
        'student': record.student,
        'rating_choices': range(1, 11),
        'back_url': (
            reverse('admin_portal:student_history', args=[record.student_id])
            if user.is_admin else
            reverse('teacher_portal:dashboard')
        ),
    })


# ---------------------------------------------------------------------------
# Edit student attendance record photo
# ---------------------------------------------------------------------------

@admin_required
def attendance_record_edit_photo(request, pk):
    """Upload or remove the daily_photo of a StudentAttendanceRecord."""
    record = get_object_or_404(
        StudentAttendanceRecord.objects.select_related('student'), pk=pk
    )
    if request.method == 'POST':
        if 'remove_photo' in request.POST:
            if record.daily_photo:
                record.daily_photo.delete(save=False)
            record.daily_photo = None
            record.save(update_fields=['daily_photo'])
            _log_audit(request, AuditLog.Action.EDIT, 'سجل حضور طالب (صورة)',
                       f'{record.student.full_name} — {record.date}')
            messages.success(request, 'تم حذف الصورة بنجاح')
        elif 'daily_photo' in request.FILES:
            if record.daily_photo:
                record.daily_photo.delete(save=False)
            record.daily_photo = request.FILES['daily_photo']
            record.save(update_fields=['daily_photo'])
            _log_audit(request, AuditLog.Action.EDIT, 'سجل حضور طالب (صورة)',
                       f'{record.student.full_name} — {record.date}')
            messages.success(request, 'تم تحديث الصورة بنجاح')
        return redirect('admin_portal:student_history', pk=record.student_id)
    return render(request, 'admin_portal/attendance_record_edit_photo.html', {
        'record': record,
        'student': record.student,
    })


# ---------------------------------------------------------------------------
# Edit teacher attendance record rating + notes (admin only)
# ---------------------------------------------------------------------------

@admin_required
def teacher_attendance_record_edit(request, pk):
    """Allow admin to set rating (1-10) and notes on a TeacherAttendanceRecord."""
    record = get_object_or_404(
        TeacherAttendanceRecord.objects.select_related('teacher'), pk=pk
    )
    back_url = reverse('admin_portal:attendance_records') + '?tab=teachers'

    if request.method == 'POST':
        notes = request.POST.get('notes', '').strip()
        try:
            rating = int(request.POST.get('rating', ''))
            if not 1 <= rating <= 10:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'التقييم يجب أن يكون رقماً من 1 إلى 10')
        else:
            record.rating = rating
            record.notes = notes
            record.save(update_fields=['rating', 'notes'])
            _log_audit(request, AuditLog.Action.EDIT, 'سجل حضور معلم',
                       f'{record.teacher.full_name} — {record.date}')
            messages.success(request, f'تم تحديث التقييم إلى {rating}/10')
            return redirect(back_url)

    return render(request, 'admin_portal/teacher_attendance_record_edit.html', {
        'record': record,
        'teacher': record.teacher,
        'rating_choices': range(1, 11),
        'back_url': back_url,
    })


# ---------------------------------------------------------------------------
# Delete attendance records (admin only)
# ---------------------------------------------------------------------------

@admin_required
@require_http_methods(['POST'])
def student_attendance_record_delete(request, pk):
    """Delete a StudentAttendanceRecord (POST only)."""
    record = get_object_or_404(
        StudentAttendanceRecord.objects.select_related('student'), pk=pk)
    _log_audit(request, AuditLog.Action.DELETE, 'سجل حضور طالب',
               f'{record.student.full_name} — {record.date}')
    record.delete()
    messages.success(request, 'تم حذف سجل الحضور بنجاح')
    next_url = request.POST.get('next', '').strip()
    if next_url:
        return redirect(next_url)
    return redirect('admin_portal:attendance_records')


@admin_required
@require_http_methods(['POST'])
def teacher_attendance_record_delete(request, pk):
    """Delete a TeacherAttendanceRecord (POST only)."""
    record = get_object_or_404(
        TeacherAttendanceRecord.objects.select_related('teacher'), pk=pk)
    _log_audit(request, AuditLog.Action.DELETE, 'سجل حضور معلم',
               f'{record.teacher.full_name} — {record.date}')
    record.delete()
    messages.success(request, 'تم حذف سجل الحضور بنجاح')
    next_url = request.POST.get('next', '').strip()
    if next_url:
        return redirect(next_url)
    return redirect(reverse('admin_portal:attendance_records') + '?tab=teachers')


# ---------------------------------------------------------------------------
# Mark teacher absent + reassign students to substitute teachers
# ---------------------------------------------------------------------------

@admin_required
def teacher_mark_absent(request, pk):
    """
    Let admin mark a teacher absent for a date, then reassign each of their
    students' attendance records to a substitute teacher (individually or in
    bulk).

    GET  ?date=YYYY-MM-DD  → show the assignment form for that date
    POST                   → save assignments and redirect
    """
    from datetime import date as date_type

    teacher = get_object_or_404(
        Teacher.objects.prefetch_related('student_links__student'), pk=pk
    )
    all_teachers = Teacher.objects.select_related(
        'user').exclude(pk=pk).order_by('full_name')

    # --- resolve date ---
    date_str = (request.POST.get('date')
                or request.GET.get('date', '')).strip()
    try:
        absence_date = date_type.fromisoformat(date_str)
    except ValueError:
        absence_date = localdate()

    # --- students linked to this teacher ---
    linked_students = list(
        Student.objects.filter(
            teacher_links__teacher=teacher).order_by('full_name')
    )

    # attendance records for the selected date keyed by student UUID string
    records_qs = (
        StudentAttendanceRecord.objects
        .filter(student__in=linked_students, date=absence_date)
        .select_related('student', 'assigned_teacher', 'original_teacher')
    )
    record_by_student = {str(r.student_id): r for r in records_qs}

    present_students = [s for s in linked_students if str(
        s.id) in record_by_student]
    not_present_students = [s for s in linked_students if str(
        s.id) not in record_by_student]

    if request.method == 'POST':
        updated = 0
        for student in present_students:
            sub_id = request.POST.get(f'sub_{student.id}', '').strip()
            if not sub_id:
                continue
            try:
                substitute = Teacher.objects.get(pk=sub_id)
            except (Teacher.DoesNotExist, ValueError, ValidationError):
                continue

            rec = record_by_student[str(student.id)]
            # Preserve original teacher as the absent teacher
            if not rec.original_teacher_id:
                rec.original_teacher = teacher
            rec.assigned_teacher = substitute
            if not rec.substitute_note:
                rec.substitute_note = f'غياب المعلم {teacher.full_name}'
            rec.save(update_fields=['original_teacher',
                     'assigned_teacher', 'substitute_note'])
            updated += 1

        if updated:
            messages.success(request, f'تم تحديث تكليف {updated} طالب بنجاح')
        else:
            messages.info(request, 'لم يتم تغيير أي تكليف')
        return redirect(reverse('admin_portal:teacher_list'))

    # Build a paired list for the template: (student, record_or_None)
    present_with_records = [
        (s, record_by_student[str(s.id)]) for s in present_students
    ]

    return render(request, 'admin_portal/teacher_mark_absent.html', {
        'teacher': teacher,
        'absence_date': absence_date,
        'all_teachers': all_teachers,
        'present_with_records': present_with_records,
        'not_present_students': not_present_students,
    })


# ---------------------------------------------------------------------------
# Audit log (admin only)
# ---------------------------------------------------------------------------

@admin_required
def audit_log(request):
    """Display all delete and edit audit log entries, newest first."""
    action_filter = request.GET.get('action', '').strip()
    object_type_filter = request.GET.get('object_type', '').strip()
    actor_filter = request.GET.get('actor', '').strip()

    qs = AuditLog.objects.all()
    if action_filter:
        qs = qs.filter(action=action_filter)
    if object_type_filter:
        qs = qs.filter(object_type=object_type_filter)
    if actor_filter:
        qs = qs.filter(actor_phone__icontains=actor_filter)

    object_types = (
        AuditLog.objects.values_list('object_type', flat=True)
        .distinct().order_by('object_type')
    )

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_portal/audit_log.html', {
        'page_obj': page_obj,
        'action_filter': action_filter,
        'object_type_filter': object_type_filter,
        'actor_filter': actor_filter,
        'object_types': object_types,
        'total_count': qs.count(),
    })
