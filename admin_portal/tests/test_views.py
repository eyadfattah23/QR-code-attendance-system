from io import BytesIO

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, Client
from django.urls import reverse
from django.utils.timezone import localdate, localtime

from core.models import User, Student, Teacher, StudentTeacherLink
from attendance.models import StudentAttendanceRecord, TeacherAttendanceRecord


def _make_excel_upload(rows):
    """Build a SimpleUploadedFile from a list of row tuples (openpyxl)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile(
        'students.xlsx',
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class AdminDashboardTestCase(TestCase):
    """Test cases for the admin dashboard view."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01234567890', email='admin@test.com', password='adminpass123',
            role=User.Role.ADMIN, first_name='Test', last_name='Admin'
        )
        cls.teacher_user = User.objects.create_user(
            phone='01234567891', email='teacher@test.com', password='teacherpass123',
            role=User.Role.TEACHER, first_name='Test', last_name='Teacher'
        )
        cls.teacher_obj = Teacher.objects.create(
            user=cls.teacher_user, full_name='Test Teacher'
        )
        cls.student1 = Student.objects.create(
            full_name='Student One', national_id='12345678901234', student_code='STU001'
        )
        cls.student2 = Student.objects.create(
            full_name='Student Two', national_id='12345678901235', student_code='STU002'
        )

    def setUp(self):
        self.client = Client()
        self.url = reverse('admin_portal:dashboard')

    # ---------- access control ----------

    def test_admin_can_access_dashboard(self):
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unauthenticated_redirected_to_login(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_teacher_cannot_access_admin_dashboard(self):
        self.client.login(phone='01234567891', password='teacherpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    # ---------- context totals ----------

    def test_context_total_students_and_teachers(self):
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['total_students'], 2)
        self.assertEqual(response.context['total_teachers'], 1)

    def test_context_today_attendance_counts(self):
        StudentAttendanceRecord.objects.create(
            student=self.student1, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user,
        )
        TeacherAttendanceRecord.objects.create(
            teacher=self.teacher_obj, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user,
        )
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_student_attendance_count'], 1)
        self.assertEqual(response.context['today_teacher_attendance_count'], 1)

    def test_context_zero_when_no_attendance_today(self):
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_student_attendance_count'], 0)
        self.assertEqual(response.context['today_teacher_attendance_count'], 0)

    # ---------- missing photos ----------

    def test_missing_photos_counts_null_records(self):
        """Records created without a photo (daily_photo=NULL) are counted."""
        StudentAttendanceRecord.objects.create(
            student=self.student1, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user, daily_photo=None,
        )
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_missing_photos_count'], 1)

    def test_missing_photos_counts_empty_string_records(self):
        """Records where photo was cleared (daily_photo='') are also counted."""
        StudentAttendanceRecord.objects.create(
            student=self.student1, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user, daily_photo='',
        )
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_missing_photos_count'], 1)

    # ---------- substitute count ----------

    def test_substitute_count_when_assigned_differs_from_original(self):
        other_teacher_user = User.objects.create_user(
            phone='01234567892', email='t2@test.com', password='t2pass',
            role=User.Role.TEACHER, first_name='Other', last_name='Teacher'
        )
        other_teacher = Teacher.objects.create(
            user=other_teacher_user, full_name='Other Teacher'
        )
        StudentAttendanceRecord.objects.create(
            student=self.student1, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user,
            original_teacher=self.teacher_obj,
            assigned_teacher=other_teacher,
        )
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_substitute_count'], 1)

    def test_substitute_count_zero_when_same_teacher(self):
        StudentAttendanceRecord.objects.create(
            student=self.student1, date=localdate(), check_in_time=localtime(),
            recorded_by=self.admin_user,
            original_teacher=self.teacher_obj,
            assigned_teacher=self.teacher_obj,
        )
        self.client.login(phone='01234567890', password='adminpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.context['today_substitute_count'], 0)


# ---------------------------------------------------------------------------
# Student management tests
# ---------------------------------------------------------------------------

class _StudentManagementBase(TestCase):
    """Shared fixtures for student management test cases."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01234567890', email='admin@test.com', password='adminpass123',
            role=User.Role.ADMIN, first_name='Test', last_name='Admin'
        )
        cls.teacher_user = User.objects.create_user(
            phone='01234567891', email='teacher@test.com', password='teacherpass123',
            role=User.Role.TEACHER, first_name='Test', last_name='Teacher'
        )
        cls.student = Student.objects.create(
            full_name='Existing Student',
            national_id='12345678901234',
            student_code='STU001',
            grade='الصف الأول',
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01234567890', password='adminpass123')


class StudentListTestCase(_StudentManagementBase):
    """Tests for student_list view."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # Extra students for pagination (25/page → need >25 total)
        for i in range(30):
            Student.objects.create(
                full_name=f'Bulk Student {i:02d}',
                national_id=f'{i:014d}',
                student_code=f'BLK{i:04d}',
            )
        cls.grade_b_student = Student.objects.create(
            full_name='Grade B Student',
            national_id='99999999999999',
            student_code='GRDB001',
            grade='الصف الثاني',
        )

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:student_list')

    def test_admin_can_access_list(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access_list(self):
        self.client.logout()
        self.client.login(phone='01234567891', password='teacherpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_search_by_name(self):
        response = self.client.get(self.url, {'q': 'Existing Student'})
        students = list(response.context['page_obj'])
        self.assertEqual(len(students), 1)
        self.assertEqual(students[0].full_name, 'Existing Student')

    def test_search_by_national_id(self):
        response = self.client.get(self.url, {'q': '12345678901234'})
        students = list(response.context['page_obj'])
        self.assertEqual(len(students), 1)

    def test_search_by_student_code(self):
        response = self.client.get(self.url, {'q': 'STU001'})
        students = list(response.context['page_obj'])
        self.assertEqual(len(students), 1)

    def test_filter_by_grade(self):
        response = self.client.get(self.url, {'grade': 'الصف الثاني'})
        students = list(response.context['page_obj'])
        self.assertEqual(len(students), 1)
        self.assertEqual(students[0].full_name, 'Grade B Student')

    def test_first_page_has_25_students(self):
        response = self.client.get(self.url)
        self.assertEqual(len(response.context['page_obj'].object_list), 25)

    def test_total_count_is_correct(self):
        response = self.client.get(self.url)
        # 1 existing + 30 bulk + 1 grade_b = 32
        self.assertEqual(response.context['total_count'], 32)


class StudentCreateTestCase(_StudentManagementBase):
    """Tests for student_create view."""

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:student_create')

    def test_get_shows_form(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('form', response.context)

    def test_create_valid_student(self):
        response = self.client.post(self.url, {
            'full_name': 'New Student',
            'national_id': '99999999999999',
            'student_code': 'NEW001',
            'grade': 'الصف الثاني',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Student.objects.filter(
            national_id='99999999999999').exists())

    def test_create_redirects_to_list(self):
        response = self.client.post(self.url, {
            'full_name': 'Another Student',
            'national_id': '88888888888888',
            'student_code': '',
            'grade': '',
        })
        self.assertRedirects(response, reverse('admin_portal:student_list'))

    def test_create_duplicate_national_id_rejected(self):
        response = self.client.post(self.url, {
            'full_name': 'Duplicate',
            'national_id': '12345678901234',  # already exists
            'student_code': 'DUP001',
            'grade': '',
        })
        # form re-rendered with errors
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Student.objects.filter(
            student_code='DUP001').exists())

    def test_create_auto_fills_student_code_from_national_id(self):
        self.client.post(self.url, {
            'full_name': 'Auto Code Student',
            'national_id': '77777777777777',
            'student_code': '',
            'grade': '',
        })
        student = Student.objects.get(national_id='77777777777777')
        self.assertEqual(student.student_code, '77777777777777')

    def test_teacher_cannot_access_create(self):
        self.client.logout()
        self.client.login(phone='01234567891', password='teacherpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)


class StudentEditTestCase(_StudentManagementBase):
    """Tests for student_edit view."""

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:student_edit', args=[self.student.id])

    def test_get_shows_form_populated(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['form'].instance.pk, self.student.pk)

    def test_edit_updates_student(self):
        self.client.post(self.url, {
            'full_name': 'Updated Name',
            'national_id': '12345678901234',
            'student_code': 'STU001',
            'grade': 'الصف الثاني',
        })
        self.student.refresh_from_db()
        self.assertEqual(self.student.full_name, 'Updated Name')
        self.assertEqual(self.student.grade, 'الصف الثاني')

    def test_edit_redirects_to_list(self):
        response = self.client.post(self.url, {
            'full_name': 'Updated Name',
            'national_id': '12345678901234',
            'student_code': 'STU001',
            'grade': '',
        })
        self.assertRedirects(response, reverse('admin_portal:student_list'))

    def test_edit_nonexistent_student_returns_404(self):
        import uuid
        url = reverse('admin_portal:student_edit', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class StudentDeleteTestCase(_StudentManagementBase):
    """Tests for student_delete view."""

    def test_delete_removes_student(self):
        new_student = Student.objects.create(
            full_name='To Delete',
            national_id='55555555555555',
            student_code='DEL001',
        )
        url = reverse('admin_portal:student_delete', args=[new_student.id])
        response = self.client.post(url)
        self.assertRedirects(response, reverse('admin_portal:student_list'))
        self.assertFalse(Student.objects.filter(pk=new_student.pk).exists())

    def test_delete_nonexistent_returns_404(self):
        import uuid
        url = reverse('admin_portal:student_delete', args=[uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)

    def test_delete_requires_post(self):
        url = reverse('admin_portal:student_delete', args=[self.student.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 405)


class StudentImportTestCase(_StudentManagementBase):
    """Tests for student_import and student_import_template views."""

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:student_import')

    def test_import_valid_excel_creates_students(self):
        upload = _make_excel_upload([
            ('full_name', 'national_id', 'student_code', 'grade'),
            ('Import Student One', '11111111111111', 'IMP001', 'الصف الثالث'),
            ('Import Student Two', '22222222222222', 'IMP002', ''),
        ])
        self.client.post(self.url, {'excel_file': upload})
        self.assertTrue(Student.objects.filter(
            national_id='11111111111111').exists())
        self.assertTrue(Student.objects.filter(
            national_id='22222222222222').exists())

    def test_import_duplicate_national_id_skipped(self):
        before = Student.objects.count()
        upload = _make_excel_upload([
            ('full_name', 'national_id'),
            ('Duplicate', '12345678901234'),  # already exists
        ])
        self.client.post(self.url, {'excel_file': upload})
        self.assertEqual(Student.objects.count(), before)

    def test_import_missing_required_headers_rejected(self):
        upload = _make_excel_upload([
            ('name', 'id'),  # wrong header names
            ('Someone', '11111111111111'),
        ])
        response = self.client.post(self.url, {'excel_file': upload})
        self.assertRedirects(response, reverse('admin_portal:student_list'))
        self.assertFalse(Student.objects.filter(
            national_id='11111111111111').exists())

    def test_import_no_file_redirects_with_error(self):
        response = self.client.post(self.url, {})
        self.assertRedirects(response, reverse('admin_portal:student_list'))

    def test_import_row_missing_name_produces_error_message(self):
        upload = _make_excel_upload([
            ('full_name', 'national_id'),
            ('', '33333333333333'),  # missing name
        ])
        response = self.client.post(self.url, {'excel_file': upload})
        self.assertFalse(Student.objects.filter(
            national_id='33333333333333').exists())
        from django.contrib.messages import get_messages
        msgs = [str(m) for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('خطأ' in m for m in msgs))

    def test_import_template_download(self):
        response = self.client.get(
            reverse('admin_portal:student_import_template'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', response['Content-Disposition'])

    def test_teacher_cannot_import(self):
        self.client.logout()
        self.client.login(phone='01234567891', password='teacherpass123')
        upload = _make_excel_upload([
            ('full_name', 'national_id'),
            ('Blocked', '44444444444444'),
        ])
        self.client.post(self.url, {'excel_file': upload})
        self.assertFalse(Student.objects.filter(
            national_id='44444444444444').exists())


# ---------------------------------------------------------------------------
# Teacher management tests
# ---------------------------------------------------------------------------

class _TeacherManagementBase(TestCase):
    """Shared fixtures for teacher management test cases."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01000000000', email='admin@test.com', password='adminpass123',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01111111111', email='teacher@test.com', password='teacherpass',
            role=User.Role.TEACHER, first_name='جورج', last_name='حبيب',
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='جورج حبيب', subject='رياضيات',
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01000000000', password='adminpass123')


class TeacherListTestCase(_TeacherManagementBase):

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:teacher_list')

    def test_admin_can_access_list(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access_list(self):
        self.client.logout()
        self.client.login(phone='01111111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_search_by_name(self):
        response = self.client.get(self.url, {'q': 'جورج'})
        teachers = list(response.context['page_obj'])
        self.assertEqual(len(teachers), 1)
        self.assertEqual(teachers[0].full_name, 'جورج حبيب')

    def test_search_by_subject(self):
        response = self.client.get(self.url, {'q': 'رياضيات'})
        self.assertEqual(len(list(response.context['page_obj'])), 1)

    def test_search_by_phone(self):
        response = self.client.get(self.url, {'q': '01111111111'})
        self.assertEqual(len(list(response.context['page_obj'])), 1)

    def test_search_no_match_returns_empty(self):
        response = self.client.get(self.url, {'q': 'لا يوجد'})
        self.assertEqual(len(list(response.context['page_obj'])), 0)

    def test_total_count_in_context(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['total_count'], 1)


class TeacherCreateTestCase(_TeacherManagementBase):

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:teacher_create')

    def test_get_shows_form(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('form', response.context)

    def test_create_valid_teacher(self):
        response = self.client.post(self.url, {
            'full_name': 'محمد سليمان',
            'subject': 'علوم',
            'phone': '01222222222',
            'first_name': 'محمد',
            'last_name': 'سليمان',
            'password': 'StrongPass1!',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Teacher.objects.filter(
            full_name='محمد سليمان').exists())
        self.assertTrue(User.objects.filter(phone='01222222222').exists())

    def test_create_redirects_to_list(self):
        response = self.client.post(self.url, {
            'full_name': 'سارة علي',
            'subject': '',
            'phone': '01333333333',
            'first_name': '',
            'last_name': '',
            'password': 'pass1234',
        })
        self.assertRedirects(response, reverse('admin_portal:teacher_list'))

    def test_create_duplicate_phone_rejected(self):
        response = self.client.post(self.url, {
            'full_name': 'مكرر',
            'subject': '',
            'phone': '01111111111',  # already in use
            'first_name': '',
            'last_name': '',
            'password': 'pass1234',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Teacher.objects.filter(full_name='مكرر').count(), 0)

    def test_create_missing_password_rejected(self):
        response = self.client.post(self.url, {
            'full_name': 'بدون كلمة مرور',
            'subject': '',
            'phone': '01444444444',
            'first_name': '',
            'last_name': '',
            'password': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(User.objects.filter(phone='01444444444').exists())

    def test_user_created_with_teacher_role(self):
        self.client.post(self.url, {
            'full_name': 'أستاذ جديد',
            'subject': 'فيزياء',
            'phone': '01555555555',
            'first_name': '',
            'last_name': '',
            'password': 'pass1234',
        })
        user = User.objects.get(phone='01555555555')
        self.assertEqual(user.role, User.Role.TEACHER)

    def test_teacher_cannot_access_create(self):
        self.client.logout()
        self.client.login(phone='01111111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)


class TeacherEditTestCase(_TeacherManagementBase):

    def setUp(self):
        super().setUp()
        self.url = reverse('admin_portal:teacher_edit', args=[self.teacher.id])

    def test_get_shows_prepopulated_form(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['form'].initial['full_name'], 'جورج حبيب')
        self.assertEqual(
            response.context['form'].initial['phone'], '01111111111')

    def test_edit_updates_teacher_fields(self):
        self.client.post(self.url, {
            'full_name': 'جورج حبيب محدث',
            'subject': 'فيزياء',
            'phone': '01111111111',
            'first_name': 'جورج',
            'last_name': 'حبيب',
            'password': '',
        })
        self.teacher.refresh_from_db()
        self.assertEqual(self.teacher.full_name, 'جورج حبيب محدث')
        self.assertEqual(self.teacher.subject, 'فيزياء')

    def test_edit_updates_user_phone(self):
        self.client.post(self.url, {
            'full_name': 'جورج حبيب',
            'subject': 'رياضيات',
            'phone': '01666666666',
            'first_name': '',
            'last_name': '',
            'password': '',
        })
        self.teacher_user.refresh_from_db()
        self.assertEqual(self.teacher_user.phone, '01666666666')

    def test_edit_blank_password_keeps_existing(self):
        import hashlib
        old_hash = self.teacher_user.password
        self.client.post(self.url, {
            'full_name': 'جورج حبيب',
            'subject': 'رياضيات',
            'phone': '01111111111',
            'first_name': '',
            'last_name': '',
            'password': '',
        })
        self.teacher_user.refresh_from_db()
        self.assertEqual(self.teacher_user.password, old_hash)

    def test_edit_with_new_password_updates_hash(self):
        self.client.post(self.url, {
            'full_name': 'جورج حبيب',
            'subject': 'رياضيات',
            'phone': '01111111111',
            'first_name': '',
            'last_name': '',
            'password': 'NewPass999',
        })
        self.teacher_user.refresh_from_db()
        self.assertTrue(self.teacher_user.check_password('NewPass999'))

    def test_edit_redirects_to_list(self):
        response = self.client.post(self.url, {
            'full_name': 'جورج حبيب',
            'subject': '',
            'phone': '01111111111',
            'first_name': '',
            'last_name': '',
            'password': '',
        })
        self.assertRedirects(response, reverse('admin_portal:teacher_list'))

    def test_edit_nonexistent_returns_404(self):
        import uuid
        url = reverse('admin_portal:teacher_edit', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class TeacherDeleteTestCase(_TeacherManagementBase):

    def test_delete_removes_teacher_and_user(self):
        new_user = User.objects.create_user(
            phone='01777777777', email='del@test.com', password='delpass',
            role=User.Role.TEACHER,
        )
        new_teacher = Teacher.objects.create(
            user=new_user, full_name='معلم للحذف',
        )
        url = reverse('admin_portal:teacher_delete', args=[new_teacher.id])
        response = self.client.post(url)
        self.assertRedirects(response, reverse('admin_portal:teacher_list'))
        self.assertFalse(Teacher.objects.filter(pk=new_teacher.pk).exists())
        self.assertFalse(User.objects.filter(pk=new_user.pk).exists())

    def test_delete_nonexistent_returns_404(self):
        import uuid
        url = reverse('admin_portal:teacher_delete', args=[uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)

    def test_delete_requires_post(self):
        url = reverse('admin_portal:teacher_delete', args=[self.teacher.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 405)


# ---------------------------------------------------------------------------
# Student-Teacher linking tests
# ---------------------------------------------------------------------------

class TeacherStudentsTestCase(TestCase):
    """Tests for the teacher_students (linking) view."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01900000000', email='admin@link.com', password='adminpass',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01911111111', email='teacher@link.com', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم الربط',
        )
        cls.student_a = Student.objects.create(
            full_name='طالب ألف', national_id='10000000000001', student_code='LNK001',
            grade='الصف الأول',
        )
        cls.student_b = Student.objects.create(
            full_name='طالب باء', national_id='10000000000002', student_code='LNK002',
            grade='الصف الثاني',
        )
        cls.student_c = Student.objects.create(
            full_name='طالب جيم', national_id='10000000000003', student_code='LNK003',
            grade='الصف الأول',
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01900000000', password='adminpass')
        self.url = reverse('admin_portal:teacher_students',
                           args=[self.teacher.id])
        # Clear links before each test
        StudentTeacherLink.objects.filter(teacher=self.teacher).delete()

    # --- access control ---

    def test_admin_can_access(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01911111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertIn('/login/', response.url)

    def test_nonexistent_teacher_returns_404(self):
        import uuid
        url = reverse('admin_portal:teacher_students', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    # --- GET context ---

    def test_get_includes_all_students(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('students', response.context)
        self.assertEqual(response.context['students'].count(), 3)

    def test_get_shows_prelinked_students_in_linked_ids(self):
        StudentTeacherLink.objects.create(
            teacher=self.teacher, student=self.student_a, is_primary=True)
        response = self.client.get(self.url)
        linked_ids = response.context['linked_ids']
        self.assertIn(str(self.student_a.id), linked_ids)
        self.assertNotIn(str(self.student_b.id), linked_ids)

    def test_get_shows_primary_ids(self):
        StudentTeacherLink.objects.create(
            teacher=self.teacher, student=self.student_a, is_primary=True)
        StudentTeacherLink.objects.create(
            teacher=self.teacher, student=self.student_b, is_primary=False)
        response = self.client.get(self.url)
        primary_ids = response.context['primary_ids']
        self.assertIn(str(self.student_a.id), primary_ids)
        self.assertNotIn(str(self.student_b.id), primary_ids)

    def test_search_filters_students(self):
        response = self.client.get(self.url, {'q': 'ألف'})
        self.assertEqual(response.context['students'].count(), 1)
        self.assertEqual(
            response.context['students'].first().full_name, 'طالب ألف')

    def test_grade_filter_filters_students(self):
        response = self.client.get(self.url, {'grade': 'الصف الأول'})
        self.assertEqual(response.context['students'].count(), 2)

    # --- POST: link ---

    def test_post_links_selected_students(self):
        self.client.post(self.url, {
            'students': [str(self.student_a.id), str(self.student_b.id)],
        })
        self.assertEqual(
            StudentTeacherLink.objects.filter(teacher=self.teacher).count(), 2)
        self.assertTrue(
            StudentTeacherLink.objects.filter(
                teacher=self.teacher, student=self.student_a).exists())

    def test_post_removes_deselected_students(self):
        # Pre-link all three
        for s in [self.student_a, self.student_b, self.student_c]:
            StudentTeacherLink.objects.create(teacher=self.teacher, student=s)
        # Submit only student_a
        self.client.post(self.url, {'students': [str(self.student_a.id)]})
        self.assertEqual(
            StudentTeacherLink.objects.filter(teacher=self.teacher).count(), 1)
        self.assertTrue(
            StudentTeacherLink.objects.filter(
                teacher=self.teacher, student=self.student_a).exists())

    def test_post_marks_primary_correctly(self):
        self.client.post(self.url, {
            'students': [str(self.student_a.id), str(self.student_b.id)],
            'primary':  [str(self.student_a.id)],
        })
        link_a = StudentTeacherLink.objects.get(
            teacher=self.teacher, student=self.student_a)
        link_b = StudentTeacherLink.objects.get(
            teacher=self.teacher, student=self.student_b)
        self.assertTrue(link_a.is_primary)
        self.assertFalse(link_b.is_primary)

    def test_post_empty_submission_removes_all_links(self):
        StudentTeacherLink.objects.create(
            teacher=self.teacher, student=self.student_a)
        self.client.post(self.url, {})  # no students submitted
        self.assertEqual(
            StudentTeacherLink.objects.filter(teacher=self.teacher).count(), 0)

    def test_post_redirects_back_to_same_page(self):
        response = self.client.post(self.url, {
            'students': [str(self.student_a.id)],
        })
        self.assertRedirects(response, self.url)

    def test_post_updates_existing_link_primary_flag(self):
        StudentTeacherLink.objects.create(
            teacher=self.teacher, student=self.student_a, is_primary=False)
        # Now mark as primary
        self.client.post(self.url, {
            'students': [str(self.student_a.id)],
            'primary':  [str(self.student_a.id)],
        })
        link = StudentTeacherLink.objects.get(
            teacher=self.teacher, student=self.student_a)
        self.assertTrue(link.is_primary)


# ---------------------------------------------------------------------------
# Student attendance history tests
# ---------------------------------------------------------------------------

class StudentHistoryTestCase(TestCase):
    """Tests for the student_history view."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01600000000', email='admin@hist.com', password='adminpass',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01611111111', email='teacher@hist.com', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم التاريخ',
        )
        cls.student = Student.objects.create(
            full_name='طالب السجل', national_id='20000000000001',
            student_code='HST001', grade='الصف الثالث',
        )
        # Two attendance records on different dates
        cls.record1 = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-01-10',
            check_in_time='2025-01-10 08:00:00',
            assigned_teacher=cls.teacher,
            rating=8,
        )
        cls.record2 = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-01-11',
            check_in_time='2025-01-11 08:05:00',
            assigned_teacher=cls.teacher,
            rating=6,
        )
        # Another student with no records
        cls.student_empty = Student.objects.create(
            full_name='طالب بلا سجل', national_id='20000000000002',
            student_code='HST002',
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01600000000', password='adminpass')
        self.url = reverse('admin_portal:student_history',
                           args=[self.student.id])

    # --- access control ---

    def test_admin_can_access(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01611111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertIn('/login/', response.url)

    def test_nonexistent_student_returns_404(self):
        import uuid
        url = reverse('admin_portal:student_history', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    # --- context ---

    def test_context_contains_student(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['student'], self.student)

    def test_context_total_records(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['total_records'], 2)

    def test_context_records_ordered_desc(self):
        response = self.client.get(self.url)
        records = list(response.context['records'])
        self.assertEqual(records[0].date.isoformat(), '2025-01-11')
        self.assertEqual(records[1].date.isoformat(), '2025-01-10')

    def test_context_teachers_present(self):
        # Link teacher to student
        StudentTeacherLink.objects.get_or_create(
            teacher=self.teacher, student=self.student)
        response = self.client.get(self.url)
        self.assertIn('teachers', response.context)

    # --- empty state ---

    def test_empty_history_student(self):
        url = reverse('admin_portal:student_history',
                      args=[self.student_empty.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_records'], 0)

    # --- template rendering ---

    def test_uses_correct_template(self):
        response = self.client.get(self.url)
        self.assertTemplateUsed(response, 'admin_portal/student_history.html')

    def test_student_name_in_response(self):
        response = self.client.get(self.url)
        self.assertContains(response, self.student.full_name)

    def test_record_dates_in_response(self):
        response = self.client.get(self.url)
        self.assertContains(response, '2025/01/10')
        self.assertContains(response, '2025/01/11')


# ---------------------------------------------------------------------------
# Attendance records browser tests
# ---------------------------------------------------------------------------

class AttendanceRecordsTestCase(TestCase):
    """Tests for the attendance_records view."""

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_user(
            phone='01700000000', email='admin@att.com', password='adminpass',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01711111111', email='teacher@att.com', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم السجلات',
        )
        cls.student_a = Student.objects.create(
            full_name='طالب أول', national_id='30000000000001',
            student_code='ATT001', grade='الصف الأول',
        )
        cls.student_b = Student.objects.create(
            full_name='طالب ثاني', national_id='30000000000002',
            student_code='ATT002', grade='الصف الثاني',
        )
        from django.utils.timezone import make_aware
        import datetime
        cls.rec_a = StudentAttendanceRecord.objects.create(
            student=cls.student_a,
            date='2025-03-01',
            check_in_time=make_aware(datetime.datetime(2025, 3, 1, 8, 0)),
            assigned_teacher=cls.teacher,
            rating=7,
        )
        cls.rec_b = StudentAttendanceRecord.objects.create(
            student=cls.student_b,
            date='2025-03-02',
            check_in_time=make_aware(datetime.datetime(2025, 3, 2, 8, 5)),
            rating=5,
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01700000000', password='adminpass')
        self.url = reverse('admin_portal:attendance_records')

    # --- access control ---

    def test_admin_can_access(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01711111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertIn('/login/', response.url)

    @classmethod
    def _make_teacher_record(cls):
        from django.utils.timezone import make_aware
        import datetime
        return TeacherAttendanceRecord.objects.create(
            teacher=cls.teacher,
            date='2025-03-05',
            check_in_time=make_aware(datetime.datetime(2025, 3, 5, 7, 45)),
        )

    # --- unfiltered context (students tab default) ---

    def test_uses_correct_template(self):
        response = self.client.get(self.url)
        self.assertTemplateUsed(
            response, 'admin_portal/attendance_records.html')

    def test_default_tab_is_students(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['tab'], 'students')

    def test_returns_all_student_records_unfiltered(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['student_total'], 2)

    def test_context_has_teachers_and_grades(self):
        response = self.client.get(self.url)
        self.assertIn('teachers', response.context)
        self.assertIn('grades', response.context)

    # --- student date filters ---

    def test_filter_date_from(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'date_from': '2025-03-02'})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_b.full_name)

    def test_filter_date_to(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'date_to': '2025-03-01'})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_a.full_name)

    def test_filter_date_range_no_results(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'date_from': '2020-01-01', 'date_to': '2020-01-31'})
        self.assertEqual(response.context['student_total'], 0)

    # --- teacher filter (on students tab) ---

    def test_filter_by_teacher(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'teacher': str(self.teacher.id)})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_a.full_name)

    # --- student search ---

    def test_filter_by_student_name(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'student': 'أول'})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_a.full_name)

    def test_filter_by_student_code(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'student': 'ATT002'})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_b.full_name)

    # --- grade filter ---

    def test_filter_by_grade(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'grade': 'الصف الأول'})
        self.assertEqual(response.context['student_total'], 1)
        self.assertContains(response, self.student_a.full_name)
        self.assertNotContains(response, self.student_b.full_name)

    # --- empty state ---

    def test_empty_student_result_renders_correctly(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'student': 'طالب غير موجود'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['student_total'], 0)

    # --- teachers tab ---

    def test_teachers_tab_loads(self):
        response = self.client.get(self.url, {'tab': 'teachers'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['tab'], 'teachers')

    def test_teachers_tab_shows_teacher_records(self):
        rec = self._make_teacher_record()
        response = self.client.get(self.url, {'tab': 'teachers'})
        self.assertEqual(response.context['teacher_total'], 1)
        self.assertContains(response, self.teacher.full_name)
        rec.delete()

    def test_teachers_tab_filter_by_name(self):
        rec = self._make_teacher_record()
        response = self.client.get(
            self.url, {'tab': 'teachers', 'teacher_q': 'السجلات'})
        self.assertEqual(response.context['teacher_total'], 1)
        response2 = self.client.get(
            self.url, {'tab': 'teachers', 'teacher_q': 'غير موجود'})
        self.assertEqual(response2.context['teacher_total'], 0)
        rec.delete()

    def test_teachers_tab_filter_by_date(self):
        rec = self._make_teacher_record()
        response = self.client.get(
            self.url, {'tab': 'teachers', 'date_from': '2025-03-06'})
        self.assertEqual(response.context['teacher_total'], 0)
        rec.delete()


# ---------------------------------------------------------------------------
# Attendance record edit rating tests
# ---------------------------------------------------------------------------

class AttendanceRecordEditRatingTestCase(TestCase):
    """Tests for the attendance_record_edit_rating view."""

    @classmethod
    def setUpTestData(cls):
        from django.utils.timezone import make_aware
        import datetime

        cls.admin_user = User.objects.create_user(
            phone='01800000000', email='admin@rating.com', password='adminpass',
            role=User.Role.ADMIN,
        )
        # Teacher linked to the student
        cls.linked_teacher_user = User.objects.create_user(
            phone='01811111111', email='linked@rating.com', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.linked_teacher = Teacher.objects.create(
            user=cls.linked_teacher_user, full_name='معلم مرتبط',
        )
        # Teacher NOT linked to the student
        cls.other_teacher_user = User.objects.create_user(
            phone='01822222222', email='other@rating.com', password='otherpass',
            role=User.Role.TEACHER,
        )
        Teacher.objects.create(user=cls.other_teacher_user,
                               full_name='معلم غير مرتبط')

        cls.student = Student.objects.create(
            full_name='طالب التقييم', national_id='40000000000001', student_code='RTG001',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.linked_teacher, student=cls.student)

        cls.record = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-04-01',
            check_in_time=make_aware(datetime.datetime(2025, 4, 1, 8, 0)),
            rating=6,
        )

    def setUp(self):
        self.client = Client()
        self.url = reverse(
            'admin_portal:attendance_record_edit_rating', args=[self.record.id])

    # --- access control ---

    def test_admin_can_access(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_linked_teacher_can_access(self):
        self.client.login(phone='01811111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unlinked_teacher_redirected(self):
        self.client.login(phone='01822222222', password='otherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        response = self.client.get(self.url)
        self.assertIn('/login/', response.url)

    # --- GET context ---

    def test_context_contains_record_and_student(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(response.context['record'], self.record)
        self.assertEqual(response.context['student'], self.student)

    def test_context_rating_choices(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(
            list(response.context['rating_choices']), list(range(1, 11)))

    # --- POST: admin updates rating ---

    def test_admin_can_update_rating(self):
        self.client.login(phone='01800000000', password='adminpass')
        self.client.post(self.url, {'rating': '9'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 9)
        self.record.rating = 6
        self.record.save(update_fields=['rating'])  # restore

    def test_admin_redirects_to_student_history(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.post(self.url, {'rating': '8'})
        self.assertRedirects(
            response,
            reverse('admin_portal:student_history', args=[self.student.id]),
        )
        self.record.rating = 6
        self.record.save(update_fields=['rating'])  # restore

    def test_linked_teacher_can_update_rating(self):
        self.client.login(phone='01811111111', password='teacherpass')
        self.client.post(self.url, {'rating': '7'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 7)
        self.record.rating = 6
        self.record.save(update_fields=['rating'])  # restore

    def test_linked_teacher_redirects_to_teacher_dashboard(self):
        self.client.login(phone='01811111111', password='teacherpass')
        response = self.client.post(self.url, {'rating': '5'})
        self.assertRedirects(response, reverse('teacher_portal:dashboard'))
        self.record.rating = 6
        self.record.save(update_fields=['rating'])  # restore

    # --- POST: invalid rating ---

    def test_invalid_rating_zero_rejected(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.post(self.url, {'rating': '0'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 6)  # unchanged
        self.assertEqual(response.status_code, 200)

    def test_invalid_rating_eleven_rejected(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.post(self.url, {'rating': '11'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 6)

    def test_invalid_rating_text_rejected(self):
        self.client.login(phone='01800000000', password='adminpass')
        response = self.client.post(self.url, {'rating': 'abc'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 6)


# ---------------------------------------------------------------------------
# Teacher attendance record edit (rating + notes) tests
# ---------------------------------------------------------------------------

class TeacherAttendanceRecordEditTestCase(TestCase):
    """Tests for the teacher_attendance_record_edit view (admin only)."""

    @classmethod
    def setUpTestData(cls):
        from django.utils.timezone import make_aware
        import datetime

        cls.admin = User.objects.create_user(
            phone='01850000000', password='adminpass',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01860000000', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم التقييم',
        )
        cls.record = TeacherAttendanceRecord.objects.create(
            teacher=cls.teacher,
            date='2025-05-01',
            check_in_time=make_aware(datetime.datetime(2025, 5, 1, 8, 0)),
            rating=7,
            notes='',
        )
        cls.url = reverse(
            'admin_portal:teacher_attendance_record_edit', args=[cls.record.pk])

    def setUp(self):
        self.client = Client()

    # --- access control ---

    def test_admin_can_access(self):
        self.client.login(phone='01850000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unauthenticated_redirected(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_teacher_cannot_access(self):
        self.client.login(phone='01860000000', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    # --- GET context ---

    def test_context_contains_record_and_teacher(self):
        self.client.login(phone='01850000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(response.context['record'], self.record)
        self.assertEqual(response.context['teacher'], self.teacher)

    def test_context_rating_choices_1_to_10(self):
        self.client.login(phone='01850000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertEqual(list(response.context['rating_choices']), list(range(1, 11)))

    def test_teacher_name_in_response(self):
        self.client.login(phone='01850000000', password='adminpass')
        response = self.client.get(self.url)
        self.assertContains(response, 'معلم التقييم')

    # --- POST: valid rating ---

    def test_admin_can_update_rating(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '9', 'notes': ''})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 9)
        self.record.rating = 7
        self.record.save(update_fields=['rating'])

    def test_admin_can_update_notes(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '7', 'notes': 'حضر متأخراً'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.notes, 'حضر متأخراً')
        self.record.notes = ''
        self.record.save(update_fields=['notes'])

    def test_admin_can_save_both_rating_and_notes(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '5', 'notes': 'ملاحظة مهمة'})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 5)
        self.assertEqual(self.record.notes, 'ملاحظة مهمة')
        self.record.rating = 7
        self.record.notes = ''
        self.record.save(update_fields=['rating', 'notes'])

    def test_successful_post_redirects_to_teachers_tab(self):
        self.client.login(phone='01850000000', password='adminpass')
        response = self.client.post(self.url, {'rating': '8', 'notes': ''})
        expected = reverse('admin_portal:attendance_records') + '?tab=teachers'
        self.assertRedirects(response, expected)
        self.record.rating = 7
        self.record.save(update_fields=['rating'])

    def test_notes_blank_on_save_clears_field(self):
        self.record.notes = 'قديم'
        self.record.save(update_fields=['notes'])
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '7', 'notes': ''})
        self.record.refresh_from_db()
        self.assertEqual(self.record.notes, '')

    # --- POST: invalid rating ---

    def test_invalid_rating_zero_rejected(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '0', 'notes': ''})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 7)

    def test_invalid_rating_eleven_rejected(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': '11', 'notes': ''})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 7)

    def test_invalid_rating_text_rejected(self):
        self.client.login(phone='01850000000', password='adminpass')
        self.client.post(self.url, {'rating': 'bad', 'notes': ''})
        self.record.refresh_from_db()
        self.assertEqual(self.record.rating, 7)

    # --- default rating value ---

    def test_new_record_default_rating_is_seven(self):
        from django.utils.timezone import make_aware
        import datetime
        new_teacher_user = User.objects.create_user(
            phone='01870000000', password='tp', role=User.Role.TEACHER)
        new_teacher = Teacher.objects.create(user=new_teacher_user, full_name='معلم جديد')
        rec = TeacherAttendanceRecord.objects.create(
            teacher=new_teacher,
            date='2025-06-01',
            check_in_time=make_aware(datetime.datetime(2025, 6, 1, 8, 0)),
        )
        self.assertEqual(rec.rating, 7)


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------

class AttendanceExportTestCase(TestCase):
    """Tests for the export_attendance_excel view."""

    @classmethod
    def setUpTestData(cls):
        from django.utils.timezone import make_aware
        import datetime

        cls.admin_user = User.objects.create_user(
            phone='01850000000', email='admin@export.com', password='adminpass',
            role=User.Role.ADMIN,
        )
        cls.teacher_user = User.objects.create_user(
            phone='01851111111', email='teacher@export.com', password='teacherpass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم التصدير',
        )
        cls.student = Student.objects.create(
            full_name='طالب التصدير', national_id='50000000000001',
            student_code='EXP001', grade='الصف الأول',
        )
        cls.rec = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-05-01',
            check_in_time=make_aware(datetime.datetime(2025, 5, 1, 8, 0)),
            assigned_teacher=cls.teacher,
            rating=8,
        )
        cls.teacher_rec = TeacherAttendanceRecord.objects.create(
            teacher=cls.teacher,
            date='2025-05-01',
            check_in_time=make_aware(datetime.datetime(2025, 5, 1, 7, 45)),
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01850000000', password='adminpass')
        self.url = reverse('admin_portal:attendance_export')

    # --- access control ---

    def test_admin_can_access(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_teacher_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01851111111', password='teacherpass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertIn('/login/', response.url)

    # --- response metadata ---

    def test_students_tab_content_type(self):
        response = self.client.get(self.url, {'tab': 'students'})
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_students_tab_filename(self):
        response = self.client.get(self.url, {'tab': 'students'})
        self.assertIn('student_attendance.xlsx',
                      response['Content-Disposition'])

    def test_teachers_tab_content_type(self):
        response = self.client.get(self.url, {'tab': 'teachers'})
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_teachers_tab_filename(self):
        response = self.client.get(self.url, {'tab': 'teachers'})
        self.assertIn('teacher_attendance.xlsx',
                      response['Content-Disposition'])

    # --- row content ---

    def _load_ws(self, response):
        import openpyxl
        import io
        return openpyxl.load_workbook(io.BytesIO(response.content)).active

    def test_students_sheet_has_header_and_data_row(self):
        response = self.client.get(self.url, {'tab': 'students'})
        ws = self._load_ws(response)
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'التاريخ')      # header
        self.assertEqual(rows[1][0], '2025-05-01')   # data row date
        self.assertEqual(rows[1][1], 'طالب التصدير')  # student name

    def test_teachers_sheet_has_header_and_data_row(self):
        response = self.client.get(self.url, {'tab': 'teachers'})
        ws = self._load_ws(response)
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'التاريخ')
        self.assertEqual(rows[1][1], 'معلم التصدير')

    # --- filters applied ---

    def test_date_filter_excludes_records(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'date_from': '2025-06-01'})
        ws = self._load_ws(response)
        rows = list(ws.iter_rows(values_only=True))
        # Only header row — no data matches future date
        self.assertEqual(len(rows), 1)

    def test_grade_filter_includes_matching(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'grade': 'الصف الأول'})
        ws = self._load_ws(response)
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)  # header + 1 data row

    def test_grade_filter_excludes_nonmatching(self):
        response = self.client.get(
            self.url, {'tab': 'students', 'grade': 'الصف الثاني'})
        ws = self._load_ws(response)
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)  # header only
