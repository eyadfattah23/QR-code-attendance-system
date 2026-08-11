import uuid

from django.test import TestCase, Client
from django.urls import reverse

from core.models import User, Student, Teacher, StudentTeacherLink
from attendance.models import StudentAttendanceRecord


class StudentHistoryTeacherPortalTestCase(TestCase):
    """Tests for teacher_portal:student_history view."""

    @classmethod
    def setUpTestData(cls):
        # Teacher who owns the student
        cls.teacher_user = User.objects.create_user(
            phone='01700000001', email='teacher1@hist.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم السجل',
        )

        # Another teacher who does NOT own the student
        cls.other_teacher_user = User.objects.create_user(
            phone='01700000002', email='teacher2@hist.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.other_teacher = Teacher.objects.create(
            user=cls.other_teacher_user, full_name='معلم آخر',
        )

        # Admin user (must NOT access teacher portal student_history)
        cls.admin_user = User.objects.create_user(
            phone='01700000003', email='admin@hist.com', password='pass',
            role=User.Role.ADMIN,
        )

        cls.student = Student.objects.create(
            full_name='طالب السجل', national_id='30000000000001',
            student_code='TH001', grade='الصف الأول',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.teacher, student=cls.student, is_primary=True,
        )

        # Unlinked student (not linked to cls.teacher)
        cls.other_student = Student.objects.create(
            full_name='طالب غير مرتبط', national_id='30000000000002',
            student_code='TH002',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.other_teacher, student=cls.other_student,
        )

        # Two attendance records on different dates
        cls.record_newer = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-03-15',
            check_in_time='2025-03-15 08:10:00',
            assigned_teacher=cls.teacher,
            original_teacher=cls.teacher,
            rating=9,
        )
        cls.record_older = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-03-10',
            check_in_time='2025-03-10 08:00:00',
            assigned_teacher=cls.teacher,
            original_teacher=cls.teacher,
            rating=7,
        )

        # A record that is a substitute assignment (different assigned/original)
        cls.substitute_record = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-03-20',
            check_in_time='2025-03-20 08:05:00',
            assigned_teacher=cls.other_teacher,
            original_teacher=cls.teacher,
            rating=8,
            substitute_note='غياب المعلم الأصلي',
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01700000001', password='pass')
        self.url = reverse('teacher_portal:student_history', args=[self.student.id])

    # --- access control ---

    def test_linked_teacher_can_access(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unlinked_teacher_gets_404(self):
        # other_student is linked to other_teacher, and has no attendance records assigned to self.teacher
        url = reverse('teacher_portal:student_history', args=[self.other_student.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_admin_cannot_access_teacher_portal(self):
        self.client.logout()
        self.client.login(phone='01700000003', password='pass')
        response = self.client.get(self.url)
        # Admin users are not teachers — should be redirected
        self.assertEqual(response.status_code, 302)

    def test_nonexistent_student_returns_404(self):
        url = reverse('teacher_portal:student_history', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    # --- context ---

    def test_context_contains_student(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['student'], self.student)

    def test_context_total_records(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['total_records'], 3)

    def test_context_records_ordered_desc_by_date(self):
        response = self.client.get(self.url)
        records = list(response.context['records'])
        dates = [r.date.isoformat() for r in records]
        self.assertEqual(dates, sorted(dates, reverse=True))

    def test_newest_record_first(self):
        response = self.client.get(self.url)
        records = list(response.context['records'])
        self.assertEqual(records[0].date.isoformat(), '2025-03-20')

    # --- template content ---

    def test_student_name_in_response(self):
        response = self.client.get(self.url)
        self.assertContains(response, 'طالب السجل')

    def test_substitute_indicator_shown(self):
        response = self.client.get(self.url)
        # The substitute record row should be highlighted (table-warning class present)
        self.assertContains(response, 'table-warning')
        # The substitute note should reference the original teacher
        self.assertContains(response, 'معلم السجل')

    def test_empty_state_when_no_records(self):
        # Create a student with no records, linked to teacher
        empty_student = Student.objects.create(
            full_name='طالب بلا سجل', national_id='30000000000099',
            student_code='TH099',
        )
        StudentTeacherLink.objects.create(teacher=self.teacher, student=empty_student)
        url = reverse('teacher_portal:student_history', args=[empty_student.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_records'], 0)
        self.assertContains(response, 'لا يوجد سجل حضور')


class EditRecordNoteTestCase(TestCase):
    """Tests for teacher_portal:edit_record_note view."""

    @classmethod
    def setUpTestData(cls):
        cls.teacher_user = User.objects.create_user(
            phone='01800000001', email='teacher_note@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم الملاحظات',
        )
        cls.other_teacher_user = User.objects.create_user(
            phone='01800000002', email='other_note@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.other_teacher = Teacher.objects.create(
            user=cls.other_teacher_user, full_name='معلم ثانٍ',
        )
        cls.admin_user = User.objects.create_user(
            phone='01800000003', email='admin_note@test.com', password='pass',
            role=User.Role.ADMIN,
        )
        cls.student = Student.objects.create(
            full_name='طالب الملاحظة', national_id='40000000000001',
            student_code='NT001',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.teacher, student=cls.student, is_primary=True,
        )
        cls.record = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-04-01',
            check_in_time='2025-04-01 08:00:00',
            assigned_teacher=cls.teacher,
            original_teacher=cls.teacher,
            rating=7,
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01800000001', password='pass')
        self.url = reverse('teacher_portal:edit_record_note', args=[self.record.id])

    # --- access control ---

    def test_linked_teacher_can_get(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unlinked_teacher_gets_404(self):
        self.client.logout()
        self.client.login(phone='01800000002', password='pass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 404)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_admin_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01800000003', password='pass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_nonexistent_record_returns_404(self):
        url = reverse('teacher_portal:edit_record_note', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    # --- GET ---

    def test_get_shows_form_with_current_note(self):
        self.record.teacher_note = 'ملاحظة قديمة'
        self.record.save(update_fields=['teacher_note'])
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ملاحظة قديمة')

    def test_context_contains_record_and_student(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['record'], self.record)
        self.assertEqual(response.context['student'], self.student)

    # --- POST save ---

    def test_post_saves_note(self):
        response = self.client.post(self.url, {'teacher_note': 'ملاحظة جديدة'})
        self.assertEqual(response.status_code, 302)
        self.record.refresh_from_db()
        self.assertEqual(self.record.teacher_note, 'ملاحظة جديدة')

    def test_post_redirects_to_dashboard(self):
        response = self.client.post(self.url, {'teacher_note': 'ملاحظة'})
        expected = reverse('teacher_portal:dashboard')
        self.assertRedirects(response, expected)

    def test_post_clears_note_when_empty(self):
        self.record.teacher_note = 'ملاحظة موجودة'
        self.record.save(update_fields=['teacher_note'])
        self.client.post(self.url, {'teacher_note': '   '})
        self.record.refresh_from_db()
        self.assertEqual(self.record.teacher_note, '')

    def test_post_unlinked_teacher_cannot_save(self):
        self.client.logout()
        self.client.login(phone='01800000002', password='pass')
        response = self.client.post(self.url, {'teacher_note': 'اختراق'})
        self.assertEqual(response.status_code, 404)
        self.record.refresh_from_db()
        self.assertNotEqual(self.record.teacher_note, 'اختراق')

    # --- template rendering ---

    def test_note_appears_in_history_after_save(self):
        self.client.post(self.url, {'teacher_note': 'ملاحظة مرئية'})
        history_url = reverse('teacher_portal:student_history', args=[self.student.id])
        response = self.client.get(history_url)
        self.assertContains(response, 'ملاحظة مرئية')


class UploadPhotoTestCase(TestCase):
    """Tests for teacher_portal:upload_photo view."""

    @classmethod
    def setUpTestData(cls):
        cls.teacher_user = User.objects.create_user(
            phone='01900000001', email='teacher_photo@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم الصور',
        )
        cls.other_teacher_user = User.objects.create_user(
            phone='01900000002', email='other_photo@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.other_teacher = Teacher.objects.create(
            user=cls.other_teacher_user, full_name='معلم آخر',
        )
        cls.admin_user = User.objects.create_user(
            phone='01900000003', email='admin_photo@test.com', password='pass',
            role=User.Role.ADMIN,
        )
        cls.student = Student.objects.create(
            full_name='طالب الصورة', national_id='50000000000001',
            student_code='PH001',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.teacher, student=cls.student, is_primary=True,
        )
        cls.record = StudentAttendanceRecord.objects.create(
            student=cls.student,
            date='2025-05-01',
            check_in_time='2025-05-01 08:00:00',
            assigned_teacher=cls.teacher,
            original_teacher=cls.teacher,
            rating=8,
        )

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01900000001', password='pass')
        self.url = reverse('teacher_portal:upload_photo', args=[self.record.id])

    def _make_jpeg(self, name='test.jpg'):
        """Return a minimal valid JPEG as SimpleUploadedFile."""
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Smallest possible valid JPEG (1×1 white pixel)
        jpeg_bytes = (
            b'\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00'
            b'\xff\xdb\x00C\x00\x08\x06\x06\x07\x06\x05\x08\x07\x07\x07\t\t'
            b'\x08\n\x0c\x14\r\x0c\x0b\x0b\x0c\x19\x12\x13\x0f\x14\x1d\x1a'
            b'\x1f\x1e\x1d\x1a\x1c\x1c $.\' ",#\x1c\x1c(7),01444\x1f\'9=82<.342\x1e1='
            b'<-6\xc0\x00\x0b\x08\x00\x01\x00\x01\x01\x01\x11\x00'
            b'\xff\xc4\x00\x1f\x00\x00\x01\x05\x01\x01\x01\x01\x01\x01\x00'
            b'\x00\x00\x00\x00\x00\x00\x00\x01\x02\x03\x04\x05\x06\x07\x08'
            b'\t\n\x0b\xff\xda\x00\x08\x01\x01\x00\x00?\x00\xfb\xd3\xff\xd9'
        )
        return SimpleUploadedFile(name, jpeg_bytes, content_type='image/jpeg')

    # --- access control ---

    def test_linked_teacher_can_get(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_unlinked_teacher_gets_404(self):
        self.client.logout()
        self.client.login(phone='01900000002', password='pass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 404)

    def test_unauthenticated_redirected(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_admin_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01900000003', password='pass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_nonexistent_record_returns_404(self):
        url = reverse('teacher_portal:upload_photo', args=[uuid.uuid4()])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    # --- GET ---

    def test_get_shows_upload_form(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'طالب الصورة')

    def test_context_contains_record_and_student(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context['record'], self.record)
        self.assertEqual(response.context['student'], self.student)

    # --- POST ---

    def test_post_no_file_shows_error(self):
        response = self.client.post(self.url, {'photo_field': 'homework_photo'})
        self.assertEqual(response.status_code, 302)
        self.record.refresh_from_db()
        self.assertFalse(bool(self.record.homework_photo))

    def test_post_invalid_type_rejected(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake_pdf = SimpleUploadedFile('doc.pdf', b'%PDF-1.4', content_type='application/pdf')
        self.client.post(self.url, {'photo_field': 'homework_photo', 'homework_photo': fake_pdf})
        self.record.refresh_from_db()
        self.assertFalse(bool(self.record.homework_photo))

    def test_post_valid_jpeg_saves_photo(self):
        photo = self._make_jpeg()
        response = self.client.post(self.url, {'photo_field': 'homework_photo', 'homework_photo': photo})
        self.assertEqual(response.status_code, 302)
        self.record.refresh_from_db()
        self.assertTrue(bool(self.record.homework_photo))

    def test_post_redirects_to_dashboard(self):
        photo = self._make_jpeg()
        response = self.client.post(self.url, {'photo_field': 'homework_photo', 'homework_photo': photo})
        self.assertRedirects(response, reverse('teacher_portal:dashboard'))

    def test_post_unlinked_teacher_cannot_upload(self):
        self.client.logout()
        self.client.login(phone='01900000002', password='pass')
        photo = self._make_jpeg()
        response = self.client.post(self.url, {'photo_field': 'homework_photo', 'homework_photo': photo})
        self.assertEqual(response.status_code, 404)
        self.record.refresh_from_db()
        self.assertFalse(bool(self.record.homework_photo))

    def tearDown(self):
        # Clean up any uploaded files to avoid leaving test artifacts
        self.record.refresh_from_db()
        if self.record.homework_photo:
            self.record.homework_photo.delete(save=True)


class ExportAttendanceTestCase(TestCase):
    """Tests for teacher_portal:export_attendance view."""

    @classmethod
    def setUpTestData(cls):
        cls.teacher_user = User.objects.create_user(
            phone='01110000001', email='export_teacher@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        cls.teacher = Teacher.objects.create(
            user=cls.teacher_user, full_name='معلم التصدير',
        )

        cls.other_teacher_user = User.objects.create_user(
            phone='01110000002', email='export_other@test.com', password='pass',
            role=User.Role.TEACHER,
        )
        # Deliberately no Teacher profile for cls.other_teacher_user

        cls.admin_user = User.objects.create_user(
            phone='01110000003', email='export_admin@test.com', password='pass',
            role=User.Role.ADMIN,
        )

        cls.student = Student.objects.create(
            full_name='طالب التصدير', national_id='40000000000001',
            student_code='EX001', grade='الصف الثاني',
        )
        StudentTeacherLink.objects.create(
            teacher=cls.teacher, student=cls.student, is_primary=True,
        )

        cls.url = reverse('teacher_portal:export_attendance')

    def setUp(self):
        self.client = Client()
        self.client.login(phone='01110000001', password='pass')

    def test_unauthenticated_redirects(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_admin_cannot_access(self):
        self.client.logout()
        self.client.login(phone='01110000003', password='pass')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def test_returns_excel_content_type(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_content_disposition_attachment(self):
        response = self.client.get(self.url)
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])

    def test_response_is_valid_xlsx(self):
        import io
        import openpyxl
        response = self.client.get(self.url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIsNotNone(wb)

    def test_student_name_in_workbook(self):
        import io
        import openpyxl
        response = self.client.get(self.url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        self.assertIn('طالب التصدير', cell_values)

    def test_attended_student_shows_present(self):
        import io
        import openpyxl
        from django.utils.timezone import localdate, now
        record = StudentAttendanceRecord.objects.create(
            student=self.student,
            date=localdate(),
            check_in_time=now(),
            recorded_by=self.teacher_user,
            original_teacher=self.teacher,
            assigned_teacher=self.teacher,
        )
        response = self.client.get(self.url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        # Row 2 is the first data row (row 1 is header)
        status_value = ws.cell(row=2, column=3).value
        self.assertEqual(status_value, 'حضر')
        record.delete()

    def test_absent_student_shows_not_recorded(self):
        import io
        import openpyxl
        response = self.client.get(self.url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        status_value = ws.cell(row=2, column=3).value
        self.assertEqual(status_value, 'لم يسجل')

    def test_teacher_with_no_profile_redirects(self):
        self.client.logout()
        self.client.login(phone='01110000002', password='pass')
        response = self.client.get(self.url)
        self.assertRedirects(response, reverse('teacher_portal:dashboard'))

    def test_header_row_columns(self):
        import io
        import openpyxl
        response = self.client.get(self.url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        self.assertIn('اسم الطالب', headers)
        self.assertIn('حالة الحضور', headers)
        self.assertIn('التقييم', headers)

